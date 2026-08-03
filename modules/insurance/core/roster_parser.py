# -*- coding: utf-8 -*-
"""花名册解析模块 - 从Excel/CSV表格或OCR文本中提取人员列表（序号+姓名）"""
import re
import os


def parse_roster_from_table(file_path):
    """
    从Excel/CSV表格文件中解析花名册，提取有序的姓名列表

    支持的表格格式：
    - Excel (.xlsx, .xls) — 使用openpyxl读取
    - CSV (.csv) — 使用csv模块读取

    表格中自动识别"姓名"列，若有序号列则按序号排序，否则按行顺序编号。

    Returns:
        list of dict: [{'seq': 1, 'name': '张三', 'idcard': '...'}, ...]
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ('.xlsx', '.xls'):
        return _parse_excel(file_path)
    elif ext == '.csv':
        return _parse_csv(file_path)
    else:
        raise ValueError(f'不支持的花名册文件格式: {ext}（请上传 .xlsx 或 .csv 文件）')


def _parse_excel(file_path):
    """解析Excel文件"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _extract_roster_from_rows(rows)


def _parse_csv(file_path):
    """解析CSV文件（自动尝试不同编码）"""
    import csv
    for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'gb18030']:
        try:
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                reader = csv.reader(f)
                rows = [list(r) for r in reader]
            if rows:
                return _extract_roster_from_rows(rows)
        except (UnicodeDecodeError, csv.Error):
            continue
    raise ValueError('CSV文件编码无法识别，请保存为UTF-8或GBK编码')


def _extract_roster_from_rows(rows):
    """
    从二维表格数据中提取花名册

    自动识别表头，找到"姓名"列，可选识别"序号"列、"身份证号"列、"人员身份类型"列。
    """
    if not rows:
        return []

    # 找到表头行（包含"姓名"关键词的行）
    header_row_idx = -1
    name_col_idx = -1
    seq_col_idx = -1
    idcard_col_idx = -1
    identity_type_col_idx = -1

    name_keywords = ['姓名', '名字', '员工姓名', '人员姓名']
    seq_keywords = ['序号', '编号', '序', '行号', 'No', 'no', 'NO']
    idcard_keywords = ['身份证', '身份证号', '证件号码', '身份证号码', '身份证明']
    identity_type_keywords = ['人员身份类型', '身份类型', '人员类别', '身份', '人员类型']

    for row_idx, row in enumerate(rows[:10]):  # 只看前10行找表头
        row_strs = [str(c).strip() if c is not None else '' for c in row]
        for col_idx, cell_val in enumerate(row_strs):
            for kw in name_keywords:
                if kw in cell_val:
                    name_col_idx = col_idx
                    header_row_idx = row_idx
                    break
            if name_col_idx >= 0:
                break
        if name_col_idx >= 0:
            # 找到了姓名列，在同一行找序号、身份证号和人员身份类型列
            for col_idx, cell_val in enumerate(row_strs):
                for kw in seq_keywords:
                    if kw in cell_val:
                        seq_col_idx = col_idx
                        break
                for kw in idcard_keywords:
                    if kw in cell_val:
                        idcard_col_idx = col_idx
                        break
                for kw in identity_type_keywords:
                    if kw in cell_val:
                        identity_type_col_idx = col_idx
                        break
            break

    # 如果没找到表头，尝试猜测第一列是序号、第二列是姓名
    if header_row_idx < 0 or name_col_idx < 0:
        # 尝试猜测：第一行是否有数据（非表头），第一列姓名
        header_row_idx = -1  # 无表头
        name_col_idx = 0
        for col_idx in range(min(5, len(rows[0]) if rows else 0)):
            # 看第一行该列是否是中文姓名
            first_val = str(rows[0][col_idx]).strip() if rows[0] and col_idx < len(rows[0]) and rows[0][col_idx] else ''
            if re.match(r'^[\u4e00-\u9fa5]{2,4}$', first_val):
                name_col_idx = col_idx
                break

    roster = []
    seen_names = set()
    data_start = header_row_idx + 1 if header_row_idx >= 0 else 0

    for row in rows[data_start:]:
        row = list(row) if not isinstance(row, list) else row
        if name_col_idx >= len(row):
            continue

        name_val = str(row[name_col_idx]).strip() if row[name_col_idx] else ''
        if not name_val or name_val.lower() in ('none', 'nan', ''):
            continue

        # 跳过非姓名数据（纯数字、空白等）
        if not re.search(r'[\u4e00-\u9fa5]', name_val):
            continue

        # 跳过表头关键词
        if _is_header_word(name_val):
            continue

        # 提取身份证号（如果有）
        idcard = ''
        if idcard_col_idx >= 0 and idcard_col_idx < len(row):
            idcard = str(row[idcard_col_idx]).strip() if row[idcard_col_idx] else ''
            # 清理可能的浮点格式
            if idcard.endswith('.0'):
                idcard = idcard[:-2]
            # 过滤非数字非X的值
            if not re.match(r'^\d{15,18}[\dXx]?$', idcard):
                idcard = ''

        # 提取序号（如果有）
        seq = None
        if seq_col_idx >= 0 and seq_col_idx < len(row):
            seq_val = row[seq_col_idx]
            if seq_val is not None:
                try:
                    seq = int(float(str(seq_val).strip()))
                except (ValueError, TypeError):
                    seq = None

        # 提取人员身份类型（如果有）
        identity_type = ''
        if identity_type_col_idx >= 0 and identity_type_col_idx < len(row):
            identity_type = str(row[identity_type_col_idx]).strip() if row[identity_type_col_idx] else ''

        if name_val not in seen_names:
            roster.append({
                'seq': seq if seq else len(roster) + 1,
                'name': name_val,
                'idcard': idcard,
                'identity_type': identity_type,
            })
            seen_names.add(name_val)

    # 按序号排序后重新编号
    roster.sort(key=lambda x: x['seq'])
    for i, item in enumerate(roster):
        item['seq'] = i + 1

    return roster


def parse_roster(text):
    """
    从OCR文本中解析花名册，提取有序的姓名列表（兼容旧版图片OCR模式）

    Returns:
        list of dict: [{'seq': 1, 'name': '张三'}, {'seq': 2, 'name': '李四'}, ...]
    """
    lines = text.split('\n')
    roster = []
    seen_names = set()

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # 跳过表头行
        header_keywords = ['序号', '姓名', '身份证', '证件号', '性别', '部门',
                           '岗位', '备注', '电话', '地址', '入职', '工号']
        if any(kw in line for kw in header_keywords) and not re.match(r'^\d{1,3}\s*[\u4e00-\u9fa5]', line):
            if re.match(r'^[^\d]*序号', line) or re.match(r'^[^\d]*姓名', line):
                continue

        # 策略1：行首数字 + 空格/制表符 + 中文名文
        m = re.match(r'^(\d{1,3})\s+([\u4e00-\u9fa5]{2,4})\b', line)
        if m:
            seq = int(m.group(1))
            name = m.group(2)
            if name not in seen_names and not _is_header_word(name):
                roster.append({'seq': seq, 'name': name})
                seen_names.add(name)
                continue

        # 策略2：行首数字紧贴中文
        m = re.match(r'^(\d{1,3})([\u4e00-\u9fa5]{2,4})', line)
        if m:
            seq = int(m.group(1))
            name = m.group(2)
            if name not in seen_names and not _is_header_word(name):
                roster.append({'seq': seq, 'name': name})
                seen_names.add(name)
                continue

        # 策略3：行中间有 "数字 中文" 模式
        m = re.search(r'(\d{1,3})\s{2,}([\u4e00-\u9fa5]{2,4})', line)
        if m:
            seq = int(m.group(1))
            name = m.group(2)
            if name not in seen_names and not _is_header_word(name):
                roster.append({'seq': seq, 'name': name})
                seen_names.add(name)
                continue

        # 策略4：行首就是纯中文姓名（没有序号，按出现顺序编号）
        m = re.match(r'^([\u4e00-\u9fa5]{2,4})\s+\d{17,18}', line)
        if m:
            name = m.group(1)
            if name not in seen_names and not _is_header_word(name):
                roster.append({'seq': len(roster) + 1, 'name': name})
                seen_names.add(name)
                continue

    # 如果没有提取到带序号的，尝试纯姓名行
    if not roster:
        for line in lines:
            line = line.strip()
            m = re.match(r'^([\u4e00-\u9fa5]{2,4})$', line)
            if m:
                name = m.group(1)
                if name not in seen_names and not _is_header_word(name):
                    roster.append({'seq': len(roster) + 1, 'name': name})
                    seen_names.add(name)

    # 按序号排序
    roster.sort(key=lambda x: x['seq'])

    # 重新编号（确保连续）
    for i, item in enumerate(roster):
        item['seq'] = i + 1

    return roster


def _is_header_word(word):
    """判断是否是表头词汇而非人名"""
    header_words = {'姓名', '性别', '民族', '出生', '日期', '住址',
                    '号码', '证件', '身份', '序号', '备注', '电话',
                    '手机', '地址', '部门', '岗位', '职务', '工号',
                    '入职', '离职', '状态', '单位', '公司', '保险',
                    '养老', '医疗', '工伤', '失业', '生育', '时间',
                    '起止', '月数', '重叠', '编号', '类型', '险种',
                    '个人', '单位', '缴费', '合计', '总计', '金额'}
    return word in header_words


def extract_roster_company_name(file_path):
    """
    从花名册文件中提取公司名称

    策略：
    1. Excel文件：读取第一个sheet的名称和第一行标题，提取公司名
    2. 常见模式："XXXX公司人员花名册"、"XXXX有限公司"等

    Returns:
        str: 公司名称，未找到返回空字符串
    """
    import re
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ('.xlsx', '.xls'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # 策略1: 检查Sheet名称
            sheet_name = ws.title.strip()
            m = re.search(r'([\u4e00-\u9fa5\w\(\)（）·]{2,30}(?:公司|企业|集团|厂|店|中心|事务所))', sheet_name)
            if m:
                wb.close()
                return m.group(1)

            # 策略2: 检查第一行（可能是标题行，如"XXXX公司人员花名册"）
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                for cell in row:
                    if cell:
                        cell_str = str(cell).strip()
                        m = re.search(r'([\u4e00-\u9fa5\w\(\)（）·]{2,30}(?:公司|企业|集团|厂|店|中心|事务所))', cell_str)
                        if m:
                            wb.close()
                            return m.group(1)

            wb.close()
        except Exception:
            pass

    elif ext == '.csv':
        try:
            for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']:
                try:
                    with open(file_path, 'r', encoding=encoding, newline='') as f:
                        for _ in range(3):
                            line = f.readline()
                            if line:
                                m = re.search(r'([\u4e00-\u9fa5\w\(\)（）·]{2,30}(?:公司|企业|集团|厂|店|中心|事务所))', line)
                                if m:
                                    return m.group(1)
                    break
                except (UnicodeDecodeError):
                    continue
        except Exception:
            pass

    return ''


def match_person_to_roster(name, roster):
    """
    将OCR识别出的姓名匹配到花名册中的序号

    Returns:
        dict or None: {'seq': 序号, 'name': 花名册中的姓名} 或 None（未匹配）
    """
    if not name or not roster:
        return None

    # 精确匹配
    for item in roster:
        if item['name'] == name:
            return item

    # 模糊匹配（OCR可能有误差）
    for item in roster:
        if name in item['name'] or item['name'] in name:
            return item

    return None


def match_person_to_roster_by_idcard(idcard, name, roster):
    """
    优先按身份证号匹配花名册，身份证号为空时回退到姓名匹配

    Args:
        idcard: OCR识别出的身份证号
        name: OCR识别出的姓名
        roster: 花名册列表

    Returns:
        dict or None: 匹配到的花名册条目，包含 identity_type 等
    """
    if not roster:
        return None

    # 优先按身份证号精确匹配
    if idcard:
        for item in roster:
            if item.get('idcard', '').strip().upper() == idcard.strip().upper():
                return item

    # 回退到姓名匹配
    return match_person_to_roster(name, roster)
