# -*- coding: utf-8 -*-
"""Excel生成模块 - 按指定表头生成社保四险重叠统计表"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .stats_calculator import INSURANCE_ORDER, calc_all_stats, parse_ym


def _thin_border():
    side = Side(style='thin', color='999999')
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header(cell):
    cell.font = Font(bold=True, size=11, color='000000', name='宋体')
    cell.alignment = Alignment(horizontal='center', vertical='center',
                               wrap_text=True)
    cell.border = _thin_border()


def _style_cell(cell, center=True, bold=False):
    cell.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='center', wrap_text=True
    )
    cell.border = _thin_border()
    cell.font = Font(size=10, bold=bold, color='000000', name='宋体')


def _style_summary_label(cell):
    cell.font = Font(bold=True, size=10, color='000000', name='宋体')
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    cell.border = _thin_border()


def _style_summary_cell(cell):
    cell.font = Font(bold=True, size=10, color='000000', name='宋体')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _thin_border()


def _style_grand_total_cell(cell):
    cell.font = Font(bold=True, size=10, color='000000', name='宋体')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _thin_border()


def _fmt_period(start, end):
    if start and end:
        return f'{start}~{end}'
    return '-'


def _build_roster_index(roster):
    """
    构建花名册索引（优先用身份证号）

    Returns:
        dict: {
            'idcard_to_entry': {idcard: entry},  # 身份证号精确匹配
            'name_to_entries': {name: [entries]},  # 姓名索引（重名时多个）
            'order_idcard': {idcard: order},  # 身份证号→花名册序号（用于排序）
            'order_name': {name: order},  # 姓名→花名册序号（兜底排序）
            'entries': [all entries]  # 全部原始条目
        }
    """
    index = {
        'idcard_to_entry': {},
        'name_to_entries': {},
        'order_idcard': {},
        'order_name': {},
        'entries': [],
    }
    if not roster:
        return index

    for i, item in enumerate(roster):
        name = item.get('name', '').strip()
        idcard = item.get('idcard', '').strip()
        index['entries'].append(item)

        if idcard:
            # 同身份证号可能多次出现（合并到第一次出现的）
            if idcard not in index['idcard_to_entry']:
                index['idcard_to_entry'][idcard] = item
                index['order_idcard'][idcard] = i

        if name:
            index['name_to_entries'].setdefault(name, []).append(item)
            if name not in index['order_name']:
                index['order_name'][name] = i

    return index


def _find_in_roster(person, roster_index):
    """
    在花名册中查找人员（先按身份证号，再按姓名）

    Args:
        person: dict {'name', 'idcard'}
        roster_index: _build_roster_index() 返回的索引

    Returns:
        matching entry 或 None
    """
    if not roster_index or not roster_index['entries']:
        return None

    idcard = person.get('idcard', '').strip()
    name = person.get('name', '').strip()

    # 1. 优先按身份证号匹配
    if idcard and idcard in roster_index['idcard_to_entry']:
        return roster_index['idcard_to_entry'][idcard]

    # 2. 兜底按姓名匹配（取第一个）
    if name and name in roster_index['name_to_entries']:
        entries = roster_index['name_to_entries'][name]
        if entries:
            return entries[0]

    return None


def _roster_order_for(person, roster_index):
    """
    返回人员在花名册中的序号位置（用于按花名册顺序排序）

    优先按身份证号，否则按姓名
    """
    idcard = person.get('idcard', '').strip()
    name = person.get('name', '').strip()

    if idcard and idcard in roster_index['order_idcard']:
        return roster_index['order_idcard'][idcard]
    if name and name in roster_index['order_name']:
        return roster_index['order_name'][name]
    return 99999


def _classify_persons(person_stats, roster_index):
    """将人员匹配到花名册（先按身份证号，再按姓名），返回 (ps, identity_type) 列表"""
    result = []
    for ps in person_stats:
        entry = _find_in_roster(ps, roster_index)
        identity_type = entry.get('identity_type', '') if entry else ''
        result.append((ps, identity_type))
    return result


def _get_year_overlap_period(overlap_start, overlap_end, year):
    """计算某一年内的重叠时间段（截取到该年度范围）"""
    if not overlap_start or not overlap_end:
        return None, None

    sy, sm = parse_ym(overlap_start)
    ey, em = parse_ym(overlap_end)

    if year < sy or year > ey:
        return None, None

    start_ym = overlap_start if year == sy else f'{year}-01'
    end_ym = overlap_end if year == ey else f'{year}-12'

    return start_ym, end_ym


def _generate_yearly_ledger(year, classified, roster_index, company_name, output_dir, timestamp):
    """为指定年份生成年度台账（独立Excel文件）

    列结构（动态）：
    - 有退役士兵时（9列）：序号/姓名/身份证号/人员身份类型/退役证编号/退役时间
      /本年度参保证明时间段/申请退税月数/申请退税金额
    - 无退役士兵时（7列）：序号/姓名/身份证号/人员身份类型
      /本年度参保证明时间段/申请退税月数/申请退税金额

    汇总行：金额在月数正下方（同一列），与总台账逻辑一致
    合计总金额行：下一列也是金额（不是月数）
    """
    # 筛选当年有重叠月数的人员
    year_persons = []
    for ps, identity_type in classified:
        if ps['has_overlap'] and ps['yearly_months'].get(year, 0) > 0:
            year_persons.append((ps, identity_type))

    if not year_persons:
        return None

    # 判断当年人员中是否有退役士兵
    has_tuiwu_year = any(it == '自主就业退役士兵' for _, it in year_persons)

    wb = Workbook()
    ws = wb.active
    ws.title = f'{year}年度台账'

    RATE_TUPIN = 650
    RATE_TUIWU = 750

    if has_tuiwu_year:
        total_cols = 9
        # 列：1序号 2姓名 3身份证号 4人员身份类型 5退役证编号 6退役时间
        #     7本年度参保证明时间段 8申请退税月数 9申请退税金额
        period_col = 7
        months_col = 8
        amount_col = 9
        headers = [
            '序号', '姓名', '身份证号', '人员身份类型',
            '退役证编号/就业创业证编号', '退役时间/登记失业时间',
            '本年度参保证明时间段\n（养老+医疗+工伤+失业）',
            '申请退税月数', '申请退税金额',
        ]
        widths = [6, 10, 22, 18, 22, 20, 28, 14, 16]
    else:
        total_cols = 7
        # 列：1序号 2姓名 3身份证号 4人员身份类型
        #     5本年度参保证明时间段 6申请退税月数 7申请退税金额
        period_col = 5
        months_col = 6
        amount_col = 7
        headers = [
            '序号', '姓名', '身份证号', '人员身份类型',
            '本年度参保证明时间段\n（养老+医疗+工伤+失业）',
            '申请退税月数', '申请退税金额',
        ]
        widths = [6, 10, 22, 18, 28, 14, 16]

    # ========== 第1行：标题 ==========
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    if company_name:
        title_text = f'{company_name}\n所属期{year}年12月申报重点群体税收优惠政策年度台账'
    else:
        title_text = f'所属期{year}年12月申报重点群体税收优惠政策年度台账'
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14, name='宋体')
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 60
    for col in range(1, total_cols + 1):
        ws.cell(row=1, column=col).border = _thin_border()

    # ========== 第2行：公司名称（盖章）+ 年度 ==========
    mid_col = (total_cols // 2) + 1
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mid_col - 1)
    ws.cell(row=2, column=1, value=f'公司名称（盖章）：{company_name}')
    ws.merge_cells(start_row=2, start_column=mid_col, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=mid_col, value=f'年度：{year}年度')
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(size=11, name='宋体')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _thin_border()
    ws.row_dimensions[2].height = 30

    # ========== 第3行：表头 ==========
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12, name='宋体')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[3].height = 40

    # ========== 数据行 ==========
    row_num = 4
    for idx, (ps, identity_type) in enumerate(year_persons):
        year_start, year_end = _get_year_overlap_period(
            ps['overlap_start'], ps['overlap_end'], year
        )
        year_period = _fmt_period(year_start, year_end) if year_start and year_end else '-'
        year_months = ps['yearly_months'].get(year, 0)
        rate = RATE_TUIWU if identity_type == '自主就业退役士兵' else RATE_TUPIN

        # v1.1.34: 序号用花名册原始序号（与姓名一一对应，不修改），未匹配到花名册时用行号
        _entry = _find_in_roster(ps, roster_index)
        seq_val = (_entry.get('seq') or idx + 1) if _entry else idx + 1

        # 构建行数据
        if has_tuiwu_year:
            row = [seq_val, ps['name'], ps['idcard'], identity_type, '', '',
                   year_period, year_months, f'=H{row_num}*{rate}']
        else:
            row = [seq_val, ps['name'], ps['idcard'], identity_type,
                   year_period, year_months, f'=F{row_num}*{rate}']

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = Font(size=11, name='宋体')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _thin_border()

        ws.row_dimensions[row_num].height = 30
        row_num += 1

    # ========== 汇总行（与总台账逻辑一致：金额在月数正下方同一列） ==========
    tupin_persons = [(ps, it) for ps, it in year_persons
                     if it in ('脱贫人口', '防止返贫监测对象')]
    tuiwu_persons_list = [(ps, it) for ps, it in year_persons
                          if it == '自主就业退役士兵']

    tupin_months = sum(ps['yearly_months'].get(year, 0) for ps, _ in tupin_persons)
    tuiwu_months = sum(ps['yearly_months'].get(year, 0) for ps, _ in tuiwu_persons_list)

    # 空行分隔
    row_num += 1

    summary_types = []
    if tupin_persons:
        summary_types.append(('脱贫人口', tupin_months, RATE_TUPIN))
    if tuiwu_persons_list:
        summary_types.append(('自主就业退役士兵', tuiwu_months, RATE_TUIWU))

    for label, months, rate in summary_types:
        # 月数行：label在period_col, 月数在months_col
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if col_idx == period_col:
                cell.value = f'{label} 月数'
            elif col_idx == months_col:
                cell.value = months
            cell.font = Font(bold=True, size=11, name='宋体')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _thin_border()
            cell.fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
        row_num += 1

        # 金额行：label在period_col, 金额在months_col（与月数同一列，在月数正下方）
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if col_idx == period_col:
                cell.value = f'{label} 金额'
            elif col_idx == months_col:
                cell.value = months * rate
            cell.font = Font(bold=True, size=11, name='宋体')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _thin_border()
            cell.fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
        row_num += 1

    # 合计总金额行：label在period_col, 金额在months_col和amount_col（两列都显示总金额）
    total_amount = tupin_months * RATE_TUPIN + tuiwu_months * RATE_TUIWU

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        if col_idx == period_col:
            cell.value = '合计总金额'
        elif col_idx == months_col:
            cell.value = total_amount
        elif col_idx == amount_col:
            cell.value = total_amount
        cell.font = Font(bold=True, size=11, name='宋体', color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()
        cell.fill = PatternFill(start_color='2E5C8A', end_color='2E5C8A', fill_type='solid')

    # ========== 列宽 ==========
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = 'A4'

    # 保存为独立Excel文件（放到"年度台账"子文件夹中）
    yearly_dir = os.path.join(output_dir, '年度台账')
    os.makedirs(yearly_dir, exist_ok=True)
    filename = f'{year}年度台账_{timestamp}.xlsx'
    filepath = os.path.join(yearly_dir, filename)
    wb.save(filepath)

    return {'filename': filename, 'filepath': filepath}


def generate_excel(persons, output_path, roster=None, company_name='', year_range=None):
    person_stats, year_cols = calc_all_stats(persons, year_range=year_range)
    roster_index = _build_roster_index(roster or [])

    # ========== 分类人员并按花名册顺序排序 ==========
    classified = _classify_persons(person_stats, roster_index)

    # 按花名册中序号排序（先按身份证号，再按姓名）
    if roster_index['entries']:
        classified.sort(key=lambda x: _roster_order_for(x[0], roster_index))

    # ========== 检查是否有退役士兵 ==========
    has_tuiwu = any(identity_type == '自主就业退役士兵' for _, identity_type in classified)

    # ========== 根据是否有退役士兵动态构建表头 ==========
    if has_tuiwu:
        headers = [
            '序号',
            '姓名',
            '身份证号',
            '人员身份类型',
            '退役证编号/就业创业证编号',
            '退役时间/登记失业时间',
            '养老保险参保证明时间段',
            '医疗保险参保证明时间段',
            '工伤保险参保证明时间段',
            '失业保险参保证明时间段',
            '参保证明时间段（养老+医疗+工伤+失业）',
            '申请退税总月数',
        ] + [f'{y}年申请退税月数' for y in year_cols] + ['合计申请退税总额']
        col_identity = 'D'
        col_overlap_months = 'L'
        col_year_start = 13
        label_col_idx = 11
        total_months_col_idx = 12
        empty_prefix_cols = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10)
        widths = [6, 10, 22, 18, 22, 20, 24, 24, 24, 24, 30, 14] + [14] * len(year_cols) + [16]
    else:
        # 没有退役士兵，不显示退役证编号和退役时间列
        headers = [
            '序号',
            '姓名',
            '身份证号',
            '人员身份类型',
            '养老保险参保证明时间段',
            '医疗保险参保证明时间段',
            '工伤保险参保证明时间段',
            '失业保险参保证明时间段',
            '参保证明时间段（养老+医疗+工伤+失业）',
            '申请退税总月数',
        ] + [f'{y}年申请退税月数' for y in year_cols] + ['合计申请退税总额']
        col_identity = 'D'
        col_overlap_months = 'J'
        col_year_start = 11
        label_col_idx = 9
        total_months_col_idx = 10
        empty_prefix_cols = (1, 2, 3, 4, 5, 6, 7, 8)
        widths = [6, 10, 22, 18, 24, 24, 24, 24, 30, 14] + [14] * len(year_cols) + [16]

    total_col_count = len(headers)
    col_total_letter = get_column_letter(total_col_count)

    wb = Workbook()
    ws = wb.active
    ws.title = '四险重叠统计'

    # ========== 标题行（第1行）==========
    if company_name:
        title_text = f'{company_name}企业申报重点群体税收优惠政策总台账'
    else:
        title_text = '企业申报重点群体税收优惠政策总台账'

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col_count)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=16, color='000000', name='宋体')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_bottom = Side(style='thin', color='999999')
    for col in range(1, total_col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.border = Border(bottom=title_bottom)
    ws.row_dimensions[1].height = 55

    # ========== 表头行（第2行）==========
    ws.append(headers)
    for cell in ws[2]:
        _style_header(cell)
    ws.row_dimensions[2].height = 40

    # ========== 人员身份类型下拉验证 ==========
    dv = DataValidation(
        type='list',
        formula1='"脱贫人口,防止返贫监测对象,自主就业退役士兵"',
        allow_blank=True
    )
    dv.error = '请选择有效的人员身份类型'
    dv.errorTitle = '输入无效'
    ws.add_data_validation(dv)

    # ========== 数据行（从第3行开始）==========
    last_data_row = 2

    for idx, (ps, identity_type) in enumerate(classified):
        row_num = idx + 3
        last_data_row = row_num

        # v1.1.34: 序号用花名册原始序号（与姓名一一对应，不修改），未匹配到花名册时用行号
        _entry = _find_in_roster(ps, roster_index)
        seq_val = (_entry.get('seq') or idx + 1) if _entry else idx + 1

        # 基础信息
        row = [seq_val, ps['name'], ps['idcard'], identity_type]

        # 退役证编号和退役时间（仅在有退役士兵时显示）
        if has_tuiwu:
            row.extend(['', ''])

        # 四险时间段
        for ins_type in INSURANCE_ORDER:
            if ins_type in ps['insurances']:
                s, e = ps['insurances'][ins_type]
                row.append(_fmt_period(s, e))
            else:
                row.append('-')

        # 重叠时间段
        if ps['has_overlap']:
            row.append(_fmt_period(ps['overlap_start'], ps['overlap_end']))
            row.append(ps['overlap_months'])
        else:
            row.append('-')
            row.append(0)

        # 各年度月数
        for y in year_cols:
            row.append(ps['yearly_months'].get(y, 0))

        # 退税总额公式（使用动态列号）
        formula = (
            f'=IF(OR(D{row_num}="脱贫人口",D{row_num}="防止返贫监测对象"),'
            f'{col_overlap_months}{row_num}*650,'
            f'IF(D{row_num}="自主就业退役士兵",{col_overlap_months}{row_num}*750,""))'
        )
        row.append(formula)

        ws.append(row)
        for cell in ws[ws.max_row]:
            _style_cell(cell)
        ws.row_dimensions[row_num].height = 25

        dv.add(f'D{row_num}')

    # ========== 分类汇总数据 ==========
    type_stats = {}

    for idx, (ps, identity_type) in enumerate(classified):
        if not identity_type:
            continue
        if identity_type not in ('脱贫人口', '防止返贫监测对象', '自主就业退役士兵'):
            continue

        key = identity_type
        if key not in type_stats:
            type_stats[key] = {
                'count': 0,
                'total_months': 0,
                'yearly': {y: 0 for y in year_cols},
                'rows': [],
            }
        type_stats[key]['count'] += 1
        type_stats[key]['total_months'] += ps.get('overlap_months', 0)
        type_stats[key]['rows'].append(idx + 3)
        for y in year_cols:
            type_stats[key]['yearly'][y] += ps.get('yearly_months', {}).get(y, 0)

    # ========== 合并脱贫人口和防止返贫监测对象 ==========
    tupin_stats = None
    if '脱贫人口' in type_stats or '防止返贫监测对象' in type_stats:
        merged = type_stats.get('脱贫人口', {'count': 0, 'total_months': 0, 'yearly': {y: 0 for y in year_cols}, 'rows': []})
        merged_fz = type_stats.get('防止返贫监测对象', {'count': 0, 'total_months': 0, 'yearly': {y: 0 for y in year_cols}, 'rows': []})
        tupin_stats = {
            'count': merged['count'] + merged_fz['count'],
            'total_months': merged['total_months'] + merged_fz['total_months'],
            'yearly': {y: merged['yearly'][y] + merged_fz['yearly'][y] for y in year_cols},
            'rows': merged['rows'] + merged_fz['rows'],
            'label': '脱贫人口',
        }

    tuwu_stats = type_stats.get('自主就业退役士兵', None)
    if tuwu_stats:
        tuwu_stats['label'] = '自主就业退役士兵'

    # ========== 结束标识行：序号完结后的下一行，合并第一列和第二列 ==========
    end_row = last_data_row + 1
    ws.merge_cells(start_row=end_row, start_column=1, end_row=end_row, end_column=2)
    end_cell = ws.cell(row=end_row, column=1, value='结束标识')
    end_cell.font = Font(size=10, color='000000', name='宋体')
    end_cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, total_col_count + 1):
        ws.cell(row=end_row, column=col).border = _thin_border()
    ws.row_dimensions[end_row].height = 25

    summary_types = []
    if tupin_stats:
        summary_types.append(tupin_stats)
    if tuwu_stats:
        summary_types.append(tuwu_stats)

    RATE_TUPIN = 650
    RATE_TUIWU = 750

    # ========== 汇总行：每类型分月数/金额两行 ==========
    amount_row_nums = []
    for st in summary_types:
        label = st['label']
        rate = RATE_TUIWU if label == '自主就业退役士兵' else RATE_TUPIN

        # 月数行
        month_row = []
        for col_idx in range(1, total_col_count + 1):
            if col_idx == label_col_idx:
                month_row.append(f'{label} 月数')
            elif col_idx in empty_prefix_cols:
                month_row.append('')
            elif col_idx == total_months_col_idx:
                month_row.append(st['total_months'])
            elif col_year_start <= col_idx < total_col_count:
                year_idx = col_idx - col_year_start
                if year_idx < len(year_cols):
                    month_row.append(st['yearly'].get(year_cols[year_idx], 0))
                else:
                    month_row.append('')
            elif col_idx == total_col_count:
                month_row.append('')

        ws.append(month_row)
        for cell_idx, cell in enumerate(ws[ws.max_row], start=1):
            if cell_idx == label_col_idx:
                _style_summary_label(cell)
            else:
                _style_summary_cell(cell)

        # 金额行
        amount_row = []
        for col_idx in range(1, total_col_count + 1):
            if col_idx == label_col_idx:
                amount_row.append(f'{label} 金额')
            elif col_idx in empty_prefix_cols:
                amount_row.append('')
            elif col_idx == total_months_col_idx:
                amount_row.append(st['total_months'] * rate)
            elif col_year_start <= col_idx < total_col_count:
                year_idx = col_idx - col_year_start
                if year_idx < len(year_cols):
                    months = st['yearly'].get(year_cols[year_idx], 0)
                    amount_row.append(months * rate)
                else:
                    amount_row.append('')
            elif col_idx == total_col_count:
                if label == '脱贫人口':
                    formula = (
                        f'=SUMIF({col_identity}$3:{col_identity}${last_data_row},'
                        f'"脱贫人口",{col_total_letter}$3:{col_total_letter}${last_data_row})'
                        f'+SUMIF({col_identity}$3:{col_identity}${last_data_row},'
                        f'"防止返贫监测对象",{col_total_letter}$3:{col_total_letter}${last_data_row})'
                    )
                else:
                    formula = (
                        f'=SUMIF({col_identity}$3:{col_identity}${last_data_row},'
                        f'"{label}",{col_total_letter}$3:{col_total_letter}${last_data_row})'
                    )
                amount_row.append(formula)

        ws.append(amount_row)
        amount_row_nums.append(ws.max_row)
        for cell_idx, cell in enumerate(ws[ws.max_row], start=1):
            if cell_idx == label_col_idx:
                _style_summary_label(cell)
            else:
                _style_summary_cell(cell)

    # ========== 合计行 ==========
    if len(summary_types) > 0:
        total_row = []

        tupin_amount = (tupin_stats['total_months'] * RATE_TUPIN) if tupin_stats else 0
        tuwu_amount = (tuwu_stats['total_months'] * RATE_TUIWU) if tuwu_stats else 0
        total_amount = tupin_amount + tuwu_amount

        for col_idx in range(1, total_col_count + 1):
            if col_idx == label_col_idx:
                total_row.append('合计总金额')
            elif col_idx in empty_prefix_cols:
                total_row.append('')
            elif col_idx == total_months_col_idx:
                total_row.append(total_amount)
            elif col_year_start <= col_idx < total_col_count:
                year_idx = col_idx - col_year_start
                if year_idx < len(year_cols):
                    tupin_months = tupin_stats['yearly'].get(year_cols[year_idx], 0) if tupin_stats else 0
                    tuwu_months = tuwu_stats['yearly'].get(year_cols[year_idx], 0) if tuwu_stats else 0
                    year_total = tupin_months * RATE_TUPIN + tuwu_months * RATE_TUIWU
                    total_row.append(year_total)
                else:
                    total_row.append('')
            elif col_idx == total_col_count:
                if len(amount_row_nums) >= 2:
                    refs = ','.join(f'{col_total_letter}{rn}' for rn in amount_row_nums)
                    formula = f'=SUM({refs})'
                elif len(amount_row_nums) == 1:
                    formula = f'={col_total_letter}{amount_row_nums[0]}'
                else:
                    formula = ''
                total_row.append(formula)

        ws.append(total_row)
        for cell in ws[ws.max_row]:
            _style_grand_total_cell(cell)

    # ========== 行高：第三行开始全部25 ==========
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 25

    # ========== 列宽 ==========
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # ========== 冻结 ==========
    ws.freeze_panes = 'A3'

    # 保存总台账
    wb.save(output_path)

    # ========== 生成年度台账（每张独立Excel文件） ==========
    output_dir = os.path.dirname(output_path)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    yearly_ledgers = []
    yearly_ledger_files = []
    for year in year_cols:
        result = _generate_yearly_ledger(
            year, classified, roster_index, company_name, output_dir, timestamp
        )
        if result:
            yearly_ledgers.append(result['filename'])
            yearly_ledger_files.append(result)

    return {
        'output_path': output_path,
        'person_count': len(person_stats),
        'year_cols': year_cols,
        'headers': headers,
        'yearly_ledgers': yearly_ledgers,
        'yearly_ledger_files': yearly_ledger_files,
    }
