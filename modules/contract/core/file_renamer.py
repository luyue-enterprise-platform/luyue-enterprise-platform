"""文件重命名模块 — 按花名册对劳动合同文件智能匹配、预览确认与批量重命名

v1.1.44 两阶段流程（预览确认后再执行）:
  1. 姓名提取: 仅从图片文件自身的文件名提取候选姓名（容错空格、分隔符、序号、业务词；
     不使用文件夹名、工号、身份证号等其他来源或字段）
  2. 精确匹配: 候选姓名与花名册"姓名"字段为唯一匹配依据，不参考工号/身份证号/部门等
     其他字段；比对时忽略姓名前后空格（含全角空格）及全角/半角字符差异
  3. 生成计划: plan_renames 生成重命名计划（不落盘），分为:
       auto      自动匹配项，新文件名 = 序号-姓名-身份证号后四位（无合法身份证号时为序号-姓名）
       duplicates 重名待确认项（花名册同名员工 / 文件名命中多个姓名），列出候选人员
       unmatched 未匹配项（查无此人 / 无法识别姓名）
       roster_missing 花名册中无对应合同的人员
  4. 预览执行: execute_renames 按用户（可在预览界面手动调整过的）最终名单执行，
     同一人员多文件追加 (2)(3) 后缀；未匹配文件移入"待处理"并生成失败明细报告
  5. 重命名日志: 记录 原文件名 -> 新文件名 对照（JSON），便于追溯
  6. 可回滚:   rollback_renames 依据日志将输出目录中的文件恢复为原文件名
"""

import os
import re
import json
import shutil
import logging
import unicodedata
from datetime import datetime

logger = logging.getLogger(__name__)

# 支持的图片格式
IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff', '.tif', '.webp'}

# ===== 命名规则 =====
# v1.1.44: {seq:02d}-{姓名}-{身份证号后4位}；身份证号缺失/格式不合法时回退 {seq:02d}-{姓名}
# 同一人员有多个文件时自动追加 (2) (3) ... 序号后缀
NAME_TEMPLATE = '{seq:02d}-{name}-{idcard_tail4}'

# 待处理文件夹与报告/日志文件名（生成在输出目录中）
PENDING_DIR_NAME = '待处理'
REPORT_FILENAME = '失败明细报告.xlsx'
LOG_FILENAME = '重命名日志.json'

# 业务词与单位词（与历史版本剔除规则保持一致）
BUSINESS_WORDS = r'(劳动合同书|劳动合同|合同书|扫描件|复印件|电子版|合同|协议)'
COMPANY_WORDS = (
    r'(有限责任公司|有限公司|股份公司|机械厂|工厂|门市部|经营部|'
    r'事务所|工作室|服务中心|公司|集团|中心|车间|班组|厂|店|部)'
)


def _normalize_name(name):
    """姓名归一化（用于精确比对）：全角转半角 + 去前后空格（含全角空格）

    NFKC 归一化可将全角字母/数字/空格/标点转为半角等价形式，
    确保文件名与花名册两侧的全角/半角差异不影响比对结果。
    """
    if not name:
        return ''
    return unicodedata.normalize('NFKC', str(name)).strip()


def _clean_raw_name(raw_name):
    """清洗原始名称：去扩展名/PDF分页后缀/业务词/单位词/序号前缀，容错空格与分隔符"""
    # v1.1.42: 先做全角->半角归一化（全角空格/数字/分隔符一并转为半角），
    # 使序号剥离与姓名提取对全角字符同样生效
    name = unicodedata.normalize('NFKC', raw_name.strip())

    # 去掉扩展名
    name = re.sub(r'\.[Pp][Dd][Ff]$', '', name)

    # 去掉PDF分页后缀 (如 xxx.pdf_p1 / xxx_p2)
    name = re.sub(r'_[Pp]\d+$', '', name)

    # 剔除业务词（任意位置）
    name = re.sub(BUSINESS_WORDS, '', name)

    # 剔除单位/机构词（仅当剔除后仍能提取到姓名时才采用，避免误伤含这些字的姓名）
    stripped = re.sub(COMPANY_WORDS, '', name)
    if re.search(r'[\u4e00-\u9fa5]{2,4}', stripped):
        name = stripped

    # 去掉常见业务前缀
    name = re.sub(r'^(扫描件|合同|劳动合同|合同书|协议)[_\-\s]*', '', name)

    # 剥离开头的序号部分 (支持 01、1、001 等数字 + 分隔符 . - _ 空格)
    name = re.sub(r'^\d+[\.\-_\s]*', '', name)

    return name.strip()


def _extract_name_candidates(raw_name):
    """从原始名称中提取全部候选姓名（去重保序）

    容错处理: 空格、点、横线、下划线等分隔符隔开的多段中文各自成为候选；
    连续无法切分的长串按贪心 2-4 字切分。

    Returns:
        list[str]: 候选姓名列表（可能为空）
    """
    cleaned = _clean_raw_name(raw_name)
    if not cleaned:
        return []

    # 连续中文段（2-4字）——分隔符会自然切断
    segments = re.findall(r'[\u4e00-\u9fa5]{2,4}', cleaned)

    seen = set()
    out = []
    for s in segments:
        if s not in seen:
            seen.add(s)
            out.append(s)
    return out


def _extract_name(raw_name):
    """兼容接口：提取第一个候选姓名（旧调用方使用）"""
    candidates = _extract_name_candidates(raw_name)
    if candidates:
        return candidates[0]
    return _clean_raw_name(raw_name)


def _build_roster_index(roster):
    """建立花名册索引: 归一化姓名 -> [人员, ...]（同名员工会对应多个人员）

    v1.1.42: 姓名经 _normalize_name 归一化（全角转半角、去前后空格）后作为索引键，
    与文件名提取的候选姓名（同样归一化）比对，忽略前后空格及全角/半角差异。
    """
    index = {}
    for person in roster:
        name = _normalize_name(person.get('name') or '')
        if name:
            index.setdefault(name, []).append(person)
    return index


def _idcard_tail(person):
    """提取身份证号后四位（仅做格式校验，不做 GB 11643 校验码验证）

    校验规则: 清理浮点尾巴（Excel 读出的 123.0）后，须为 15 位纯数字
    或 18 位（前 17 位数字 + 末位数字/X）。不合法或缺失返回空串，
    命名时回退为"序号-姓名"格式。
    """
    raw = person.get('idcard') or ''
    raw = re.sub(r'\.0+$', '', str(raw).strip().upper())
    if re.fullmatch(r'\d{15}|\d{17}[\dX]', raw):
        return raw[-4:]
    return ''


def _person_brief(person):
    """人员摘要（进入 plan / 前端预览的脱敏结构，不含完整身份证号）"""
    return {
        'seq': person.get('seq'),
        'name': _normalize_name(person.get('name')),
        'idcard_tail': _idcard_tail(person),
    }


def _render_new_base(person):
    """渲染新文件名主干（不含扩展名和 (2)(3) 序号后缀）

    v1.1.44: {seq:02d}-{姓名}-{身份证号后4位}；
    身份证号缺失/格式不合法时回退 {seq:02d}-{姓名}。
    姓名经 _normalize_name 归一化，避免花名册中带前后空格的姓名进入文件名。
    """
    name = _normalize_name(person.get('name'))
    tail = _idcard_tail(person)
    try:
        seq = int(person.get('seq'))
        seq_str = f'{seq:02d}'
    except (TypeError, ValueError):
        seq_str = str(person.get('seq', ''))
    if tail:
        return f'{seq_str}-{name}-{tail}'
    return f'{seq_str}-{name}'


def _unique_name(output_dir, base, ext):
    """生成不冲突的文件名：base.ext 已存在时追加 (2) (3) ..."""
    candidate = base + ext
    counter = 2
    while os.path.exists(os.path.join(output_dir, candidate)):
        candidate = f'{base}({counter}){ext}'
        counter += 1
    return candidate


# ===== 阶段1: 生成重命名计划（不落盘） =====

def plan_renames(file_paths, roster, folder_hints=None):
    """按花名册对劳动合同文件智能匹配，生成重命名计划（预览用，不执行任何重命名）

    命名规则:
    - 身份证号合法: {seq:02d}-{姓名}-{身份证号后4位}.{ext}
    - 身份证号缺失/不合法: {seq:02d}-{姓名}.{ext}
    - 同一人员多个文件: 第2个起追加 (2) (3) ... 后缀

    Args:
        file_paths: 文件路径列表
        roster: 花名册人员列表 [{'seq': 1, 'name': '张三', 'idcard': '...'}, ...]
        folder_hints: 兼容保留参数（v1.1.42 起姓名匹配仅依据图片文件自身的文件名，
                      文件夹名不再参与匹配，此参数被忽略）

    Returns:
        dict: {
            'auto': [{original, seq, name, new_name}],   # 自动匹配（new_name 含扩展名）
            'duplicates': [{original, guessed, reason,
                            candidates: [{seq, name, idcard_tail}]}],  # 重名待确认
            'unmatched': [{original, guessed, reason}],  # 未匹配
            'roster_missing': [{seq, name}],             # 花名册中无对应合同的人员
            'total': int,
        }
    """
    roster_index = _build_roster_index(roster)

    matched_files = {}    # {seq: [(path, basename)]}
    duplicates = []       # 重名待确认
    unmatched = []        # 未匹配
    dup_names = set()     # 有重名待确认文件的姓名（roster_missing 排除用）
    matched_seqs = set()  # 已匹配（含多文件）的花名册序号

    for fp in file_paths:
        basename = os.path.basename(fp)
        file_base = os.path.splitext(basename)[0]

        # v1.1.42: 姓名仅从图片文件自身的文件名读取，文件夹名不参与匹配；
        # 候选姓名与花名册"姓名"字段（两侧均归一化）为唯一匹配依据
        candidates_names = _extract_name_candidates(file_base)
        guessed = candidates_names[0] if candidates_names else ''

        # 命中的花名册姓名（完全一致才算命中）
        hits = [c for c in candidates_names if c in roster_index]

        if len(hits) >= 2:
            # 文件名中含多个姓名 -> 重名待确认
            cand_persons = [_person_brief(p) for h in hits for p in roster_index[h]]
            duplicates.append({
                'original': basename,
                'guessed': '、'.join(hits),
                'reason': '文件名中包含多个花名册姓名，请确认归属',
                'candidates': cand_persons,
            })
            dup_names.update(hits)
            logger.warning(f'重名待确认(多姓名): {basename!r} 命中 {hits}')
        elif len(hits) == 1:
            persons = roster_index[hits[0]]
            if len(persons) > 1:
                # 花名册存在同名员工 -> 重名待确认
                duplicates.append({
                    'original': basename,
                    'guessed': hits[0],
                    'reason': '花名册中存在同名员工，请确认归属',
                    'candidates': [_person_brief(p) for p in persons],
                })
                dup_names.add(hits[0])
                logger.warning(f'重名待确认(同名): {basename!r} 花名册有 {len(persons)} 个 {hits[0]!r}')
            else:
                seq = persons[0].get('seq')
                matched_files.setdefault(seq, []).append((fp, basename))
                matched_seqs.add(seq)
        else:
            # 未命中 —— 给出明确的失败原因
            if candidates_names:
                reason = f'提取姓名"{guessed}"在花名册中查无此人，请核对文件名或补充花名册'
            else:
                reason = '文件名中无法识别出姓名，无法与花名册匹配'
            unmatched.append({'original': basename, 'guessed': guessed, 'reason': reason})
            logger.warning(f'未匹配: {basename!r} 提取={guessed!r} ({reason})')

    # 生成自动匹配项的新文件名（同一人员多文件在计划阶段即分配 (2)(3) 后缀）
    auto = []
    for person in roster:
        seq = person.get('seq')
        files = matched_files.get(seq, [])
        if not files:
            continue
        base = _render_new_base(person)
        for idx, (fp, basename) in enumerate(files):
            ext = os.path.splitext(basename)[1].lower()
            if ext not in IMAGE_EXTENSIONS:
                ext = '.jpg'  # 兜底
            if idx == 0:
                new_name = f'{base}{ext}'
            else:
                new_name = f'{base}({idx + 1}){ext}'
            auto.append({
                'original': basename,
                'seq': seq,
                'name': _normalize_name(person.get('name')),
                'new_name': new_name,
            })

    # 花名册中无对应合同的人员（排除已匹配序号；有重名待确认文件的人员等用户确认）
    roster_missing = []
    for person in roster:
        seq = person.get('seq')
        if seq in matched_seqs:
            continue
        name = _normalize_name(person.get('name') or '')
        if name and name in dup_names:
            continue
        roster_missing.append({'seq': seq, 'name': name})

    plan = {
        'auto': auto,
        'duplicates': duplicates,
        'unmatched': unmatched,
        'roster_missing': roster_missing,
        'total': len(file_paths),
    }
    logger.info(
        '重命名计划已生成: 共 %d 个文件, 自动匹配 %d, 重名待确认 %d, 未匹配 %d, 花名册无合同 %d',
        plan['total'], len(auto), len(duplicates), len(unmatched), len(roster_missing),
    )
    return plan


# ===== 阶段2: 校验与执行 =====

def validate_renames(source_paths, renames):
    """执行前同步校验用户提交的重命名名单（错误立即返回，不进线程）

    Args:
        source_paths: {原文件名: 路径} 映射（来自任务上传目录）
        renames: [{'original', 'new_name', 'seq'?}, ...]

    Returns:
        list[str]: 错误信息列表（空列表表示通过）
    """
    errors = []
    seen = {}
    for r in renames:
        original = r.get('original') or ''
        new_name = r.get('new_name') or ''
        if not original:
            errors.append('存在缺少原文件名的重命名项')
            continue
        if original not in source_paths:
            errors.append(f'原文件不存在: {original}')
            continue
        if not new_name or not new_name.strip():
            errors.append(f'{original}: 新文件名不能为空')
            continue
        # 非法字符（路径分隔符等）
        if re.search(r'[/\\]', new_name) or new_name.strip() in {'.', '..'}:
            errors.append(f'{original}: 新文件名含非法字符')
            continue
        # 查重（Windows 文件系统大小写不敏感）
        key = new_name.strip().lower()
        if key in seen and seen[key] != original:
            errors.append(f'新文件名重复: {new_name}（{seen[key]} 与 {original}）')
            continue
        seen[key] = original
    return errors


def execute_renames(source_paths, plan, output_dir, renames, pending,
                    progress_callback=None, task_id=None):
    """按用户确认（可手动调整）的最终名单执行批量重命名

    执行内容:
    1. renames 中的文件按 new_name 复制到输出目录（防冲突自动加 (2)(3)）
    2. pending + 计划未匹配文件移入"待处理"文件夹
    3. 生成失败明细报告（Excel）与重命名日志（JSON）

    Args:
        source_paths: {原文件名: 路径} 映射
        plan: plan_renames 的计划（用于回填 person 信息与未匹配清单）
        output_dir: 输出目录
        renames: [{'original', 'new_name', 'seq'?}, ...]（新文件名由前端决定，
                 重名项的 seq 用于从候选人员中回填信息）
        pending: [original, ...] 用户标记为移入待处理的原文件名
        progress_callback: 可选，callback(current, total)
        task_id: 任务ID（写入日志）

    Returns:
        dict: {
            'renamed': [{original, new_name, person_seq, person_name}],
            'unmatched': [{original, guessed, reason}],
            'total': int, 'matched_count': int, 'unmatched_count': int,
            'conflicts_resolved_count': int, 'output_dir': str,
        }
    """
    os.makedirs(output_dir, exist_ok=True)

    # 计划信息回填用：auto 项 person 信息；重名项候选人员（按 seq 索引）
    auto_by_original = {item['original']: item for item in plan.get('auto', [])}
    dup_by_original = {item['original']: item for item in plan.get('duplicates', [])}

    # 计划中的未匹配清单（与用户 pending 合并去重）
    unmatched_by_original = {
        item['original']: dict(item) for item in plan.get('unmatched', [])
    }

    renamed = []
    total = len(renames) + len(pending) + len(unmatched_by_original)
    done = 0

    for r in renames:
        original = r.get('original') or ''
        new_name = (r.get('new_name') or '').strip()
        fp = source_paths.get(original)
        if not fp or not os.path.exists(fp):
            logger.error(f'执行重命名跳过（源文件不存在）: {original}')
            continue

        # 无扩展名时补原扩展名
        if not os.path.splitext(new_name)[1]:
            new_name = new_name + os.path.splitext(original)[1].lower()

        # 防文件系统冲突兜底（同名已在 validate 拦截，此处防磁盘上既有文件）
        base, ext = os.path.splitext(new_name)
        new_name = _unique_name(output_dir, base, ext)

        new_path = os.path.join(output_dir, new_name)
        shutil.copy2(fp, new_path)

        # 回填 person 信息
        person_seq = None
        person_name = ''
        auto_item = auto_by_original.get(original)
        if auto_item:
            person_seq = auto_item.get('seq')
            person_name = auto_item.get('name', '')
        else:
            seq = r.get('seq')
            dup_item = dup_by_original.get(original)
            if seq is not None and dup_item:
                for cand in dup_item.get('candidates', []):
                    if cand.get('seq') == seq:
                        person_seq = seq
                        person_name = cand.get('name', '')
                        break

        renamed.append({
            'original': original,
            'new_name': new_name,
            'person_seq': person_seq,
            'person_name': person_name,
        })

        done += 1
        if progress_callback:
            try:
                progress_callback(done, total)
            except Exception:
                pass

    # 用户标记移入待处理 + 计划未匹配（去重合并）
    final_unmatched = []
    seen_pending = set()
    for original in list(pending or []):
        if original in seen_pending:
            continue
        seen_pending.add(original)
        item = unmatched_by_original.pop(original, None)
        if item:
            final_unmatched.append(item)
        else:
            dup_item = dup_by_original.get(original)
            final_unmatched.append({
                'original': original,
                'guessed': (dup_item or {}).get('guessed', ''),
                'reason': '人工确认：移入待处理文件夹',
            })
    final_unmatched.extend(unmatched_by_original.values())

    # 待处理文件移入"待处理"文件夹
    if final_unmatched:
        pending_dir = os.path.join(output_dir, PENDING_DIR_NAME)
        os.makedirs(pending_dir, exist_ok=True)
        for item in final_unmatched:
            fp = source_paths.get(item['original'])
            if fp and os.path.exists(fp):
                dest = os.path.join(pending_dir, item['original'])
                counter = 1
                while os.path.exists(dest):
                    stem, ext = os.path.splitext(item['original'])
                    dest = os.path.join(pending_dir, f'{stem}({counter}){ext}')
                    counter += 1
                shutil.copy2(fp, dest)

    # 重名确认数 = 提交的重命名项中不属于 auto 的数量
    conflicts_resolved = sum(1 for r in renames if r.get('original') not in auto_by_original)

    result = {
        'renamed': renamed,
        'unmatched': final_unmatched,
        'total': plan.get('total', len(renamed) + len(final_unmatched)),
        'matched_count': len(renamed),
        'unmatched_count': len(final_unmatched),
        'conflicts_resolved_count': conflicts_resolved,
        'output_dir': output_dir,
    }

    # 失败明细报告 + 重命名日志
    write_failure_report(output_dir, result['unmatched'])
    write_rename_log(output_dir, result, task_id=task_id)

    logger.info(
        '重命名执行完成: 共 %d 个文件, 重命名 %d, 待处理 %d, 重名确认 %d',
        result['total'], result['matched_count'],
        result['unmatched_count'], result['conflicts_resolved_count'],
    )
    return result


# ===== 阶段3: 回滚 =====

def rollback_renames(output_dir):
    """依据重命名日志回滚输出目录中的文件（恢复原文件名）

    回滚语义:
    - 对日志中每条 rename 记录: 若 原文件名 不存在且 新文件名 存在，
      则将 新文件名 改回 原文件名（仅在输出目录内操作，不影响待处理文件夹与源上传目录）
    - 原文件名已存在则跳过（避免覆盖）
    - 回滚记录追加写入日志的 rollbacks 字段

    Returns:
        dict: {'reverted': int, 'failed': int}
        或 {'error': str}（日志不存在/不可读时）
    """
    log_path = os.path.join(output_dir, LOG_FILENAME)
    if not os.path.exists(log_path):
        return {'error': '重命名日志不存在，无法回滚'}

    try:
        with open(log_path, 'r', encoding='utf-8') as f:
            log = json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        logger.error(f'读取重命名日志失败: {e}')
        return {'error': f'读取重命名日志失败: {e}'}

    reverted = 0
    failed = 0
    rollbacks = []

    for item in log.get('renames', []):
        original = item.get('original')
        new_name = item.get('new_name')
        if not original or not new_name:
            continue
        orig_path = os.path.join(output_dir, original)
        new_path = os.path.join(output_dir, new_name)
        if os.path.exists(orig_path):
            continue  # 原文件名已存在，跳过避免覆盖
        if not os.path.exists(new_path):
            continue  # 新文件名不存在（可能已手动处理）
        try:
            os.rename(new_path, orig_path)
            reverted += 1
            rollbacks.append({
                'time': datetime.now().isoformat(timespec='seconds'),
                'new_name': new_name,
                'restored_to': original,
            })
        except OSError as e:
            failed += 1
            logger.error(f'回滚失败 {new_name} -> {original}: {e}')

    if rollbacks:
        log.setdefault('rollbacks', []).extend(rollbacks)
        try:
            with open(log_path, 'w', encoding='utf-8') as f:
                json.dump(log, f, ensure_ascii=False, indent=2)
        except OSError as e:
            logger.error(f'回滚记录写入日志失败: {e}')

    logger.info(f'回滚完成: 恢复 {reverted} 个, 失败 {failed} 个')
    return {'reverted': reverted, 'failed': failed}


# ===== 失败明细报告（Excel） =====

def write_failure_report(output_dir, failures):
    """生成失败明细报告（Excel），写入输出目录的"待处理"文件夹

    Args:
        output_dir: 输出目录
        failures: [{original, guessed, reason, candidates?}, ...]
                  candidates: [{'seq', 'name'}] 冲突候选人员（可选）
    """
    if not failures:
        return None

    pending_dir = os.path.join(output_dir, PENDING_DIR_NAME)
    os.makedirs(pending_dir, exist_ok=True)
    report_path = os.path.join(pending_dir, REPORT_FILENAME)

    try:
        import openpyxl
    except ImportError:
        logger.error('生成失败明细报告需要 openpyxl')
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '失败明细'

    headers = ['序号', '原文件名', '提取姓名', '失败原因', '候选人员', '处理建议']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for i, item in enumerate(failures, 1):
        candidates = item.get('candidates') or []
        cand_str = '；'.join(f"{c.get('seq')}-{c.get('name')}" for c in candidates)
        if candidates:
            suggestion = '请在系统中人工确认归属后重新处理'
        elif '查无此人' in item.get('reason', '') or '不在花名册' in item.get('reason', ''):
            suggestion = '请核对文件名或补充花名册'
        else:
            suggestion = '请手动重命名后放入结果文件夹'
        ws.append([
            i,
            item.get('original', ''),
            item.get('guessed', ''),
            item.get('reason', ''),
            cand_str,
            suggestion,
        ])

    # 简单列宽
    widths = [6, 36, 12, 30, 24, 32]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(report_path)
    logger.info(f'失败明细报告已生成: {report_path} ({len(failures)} 条)')
    return report_path


# ===== 重命名日志（JSON，可回溯/回滚） =====

def write_rename_log(output_dir, result, task_id=None):
    """生成重命名日志（JSON）：原文件名->新文件名对照、失败明细、回滚记录

    Args:
        output_dir: 输出目录
        result: execute_renames 的结果 dict
        task_id: 任务ID
    """
    log_path = os.path.join(output_dir, LOG_FILENAME)

    now = datetime.now().isoformat(timespec='seconds')
    log = {
        'task_id': task_id,
        'executed_at': now,
        'naming_rule': NAME_TEMPLATE + '（无合法身份证号时回退 {seq:02d}-{name}）',
        'summary': {
            'total': result.get('total', 0),
            'renamed': result.get('matched_count', 0),
            'pending': result.get('unmatched_count', 0),
            'conflicts_resolved': result.get('conflicts_resolved_count', 0),
        },
        'renames': [
            {
                'time': now,
                'original': item['original'],
                'new_name': item['new_name'],
                'person_seq': item.get('person_seq'),
                'person_name': item.get('person_name'),
            }
            for item in result.get('renamed', [])
        ],
        'failures': [
            {
                'time': now,
                'original': item['original'],
                'guessed': item.get('guessed', ''),
                'reason': item.get('reason', ''),
            }
            for item in result.get('unmatched', [])
        ],
        'rollbacks': [],
    }

    try:
        with open(log_path, 'w', encoding='utf-8') as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
        logger.info(f'重命名日志已生成: {log_path}')
    except OSError as e:
        logger.error(f'重命名日志写入失败: {e}')

    return log_path
