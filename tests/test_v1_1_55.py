# -*- coding: utf-8 -*-
"""v1.1.55 五项需求测试

需求1：修正重叠参保时间段的统计逻辑——统计结果起点 = max(统计开始, 重叠起点)，
       终点保持重叠实际结束（不按统计截止裁剪）。
       用户给定示例（统计时间段 2023-01~2025-12）：
       - 重叠 2023-04~2026-07 → 统计结果 2023-04~2026-07
       - 重叠 2022-09~2026-07 → 统计结果 2023-01~2026-07
       - 重叠 2023-01~2026-07 → 统计结果 2023-01~2026-07
       其余计算规则不变（年度/总月数跟筛选走，v1.1.38）。
"""
import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.insurance.core.stats_calculator import (
    apply_stat_range_clamp,
    calc_all_stats,
    calc_person_stats,
)


def _make_person(name='张三', idcard='610102199001011234', period=None):
    """构造单险种统一时间段的人员（四险一致 → 重叠=该时间段）"""
    if period is None:
        period = ('2022-09', '2026-07')
    return {
        'name': name,
        'idcard': idcard,
        'insurances': {ins: period for ins in
                       ['养老保险', '医疗保险', '工伤保险', '失业保险']},
    }


class TestApplyStatRangeClampExamples(unittest.TestCase):
    """需求1：用户示例逐条验证（统计时间段 2023-01~2025-12）"""

    def _clamp_one(self, overlap_period, year_range=('2023-01', '2025-12')):
        ps = calc_person_stats(_make_person(period=overlap_period),
                               year_range=year_range)
        apply_stat_range_clamp([ps], year_range)
        return ps

    def test_example1_overlap_starts_inside_range(self):
        """重叠 2023-04~2026-07 → 统计结果 2023-04~2026-07（起点晚于统计开始，不变）"""
        ps = self._clamp_one(('2023-04', '2026-07'))
        self.assertTrue(ps['has_overlap'])
        self.assertEqual(ps['overlap_start'], '2023-04')
        self.assertEqual(ps['overlap_end'], '2026-07')

    def test_example2_overlap_starts_before_range(self):
        """重叠 2022-09~2026-07 → 统计结果 2023-01~2026-07（起点钳到统计开始）"""
        ps = self._clamp_one(('2022-09', '2026-07'))
        self.assertTrue(ps['has_overlap'])
        self.assertEqual(ps['overlap_start'], '2023-01')
        self.assertEqual(ps['overlap_end'], '2026-07')  # 终点不按统计截止裁剪

    def test_example3_overlap_equals_range_start(self):
        """重叠 2023-01~2026-07 → 统计结果 2023-01~2026-07（两者一致，不变）"""
        ps = self._clamp_one(('2023-01', '2026-07'))
        self.assertTrue(ps['has_overlap'])
        self.assertEqual(ps['overlap_start'], '2023-01')
        self.assertEqual(ps['overlap_end'], '2026-07')

    def test_months_unchanged_after_clamp(self):
        """月数规则不变：总月数仍=筛选范围内年度月数之和（v1.1.38）"""
        ps = self._clamp_one(('2022-09', '2026-07'))
        self.assertEqual(ps['overlap_months'], 36)  # 2023~2025 各 12 个月
        self.assertEqual(ps['yearly_months'].get(2023), 12)
        self.assertEqual(ps['yearly_months'].get(2024), 12)
        self.assertEqual(ps['yearly_months'].get(2025), 12)
        self.assertEqual(ps['yearly_months'].get(2026), 0)

    def test_yearly_keys_rebased_to_clamped_start(self):
        """年度键随钳制后起点重排：2022 键剔除，保留键值不变"""
        ps = self._clamp_one(('2022-09', '2026-07'))
        self.assertNotIn(2022, ps['yearly_months'])
        self.assertEqual(list(ps['yearly_months'].keys()), [2023, 2024, 2025, 2026])
        self.assertEqual(ps['years'], [2023, 2024, 2025, 2026])

    def test_partial_year_start_clamp(self):
        """统计开始为年中（2023-06）：重叠 2022-09~2026-07 → 2023-06~2026-07"""
        ps = self._clamp_one(('2022-09', '2026-07'), year_range=('2023-06', '2025-12'))
        self.assertEqual(ps['overlap_start'], '2023-06')
        self.assertEqual(ps['overlap_end'], '2026-07')
        self.assertEqual(ps['overlap_months'], 31)  # 2023:7 + 2024:12 + 2025:12


class TestApplyStatRangeClampEdges(unittest.TestCase):
    """需求1：边界情况"""

    def test_no_year_range_unchanged(self):
        """未设统计时间段 → 不钳制，保持全量重叠"""
        ps = calc_person_stats(_make_person(period=('2022-09', '2026-07')))
        apply_stat_range_clamp([ps], None)
        self.assertEqual(ps['overlap_start'], '2022-09')
        self.assertEqual(ps['overlap_end'], '2026-07')

    def test_empty_year_range_unchanged(self):
        """year_range 为空列表/空元组 → 不钳制"""
        ps = calc_person_stats(_make_person(period=('2022-09', '2026-07')))
        apply_stat_range_clamp([ps], ())
        self.assertEqual(ps['overlap_start'], '2022-09')

    def test_no_intersection_becomes_no_result(self):
        """统计开始晚于重叠终点（无交集）→ 按 无统计结果 处理"""
        ps = calc_person_stats(_make_person(period=('2022-01', '2022-12')),
                               year_range=('2023-01', '2025-12'))
        apply_stat_range_clamp([ps], ('2023-01', '2025-12'))
        self.assertFalse(ps['has_overlap'])
        self.assertIsNone(ps['overlap_start'])
        self.assertIsNone(ps['overlap_end'])
        self.assertEqual(ps['overlap_months'], 0)
        self.assertEqual(ps['years'], [])
        self.assertEqual(ps['yearly_months'], {})

    def test_overlap_entirely_after_range_kept(self):
        """重叠整体晚于统计开始 → 起点不变（max 取重叠起点）"""
        ps = calc_person_stats(_make_person(period=('2024-05', '2026-07')),
                               year_range=('2023-01', '2025-12'))
        apply_stat_range_clamp([ps], ('2023-01', '2025-12'))
        self.assertEqual(ps['overlap_start'], '2024-05')
        self.assertEqual(ps['overlap_end'], '2026-07')

    def test_no_overlap_person_untouched(self):
        """本就无重叠的人员 → 原样跳过不报错"""
        person = {
            'name': '李四', 'idcard': '610102199001015678',
            'insurances': {'养老保险': ('2022-01', '2022-12'),
                           '医疗保险': ('2024-01', '2024-12')},
        }
        ps = calc_person_stats(person, year_range=('2023-01', '2025-12'))
        self.assertFalse(ps['has_overlap'])
        apply_stat_range_clamp([ps], ('2023-01', '2025-12'))
        self.assertFalse(ps['has_overlap'])

    def test_malformed_year_range_ignored(self):
        """year_range 格式异常 → 安全跳过不抛错"""
        ps = calc_person_stats(_make_person(period=('2022-09', '2026-07')))
        apply_stat_range_clamp([ps], ('bad-01', '2025-12'))
        self.assertEqual(ps['overlap_start'], '2022-09')

    def test_mixed_population(self):
        """多人混合：仅起点早于统计开始的人被钳制，其余不动"""
        persons = [_make_person('甲', '610102199001011001', ('2022-09', '2026-07')),
                   _make_person('乙', '610102199001011002', ('2023-04', '2026-07')),
                   _make_person('丙', '610102199001011003', ('2022-01', '2022-12'))]
        stats, _ = calc_all_stats(persons, year_range=('2023-01', '2025-12'))
        apply_stat_range_clamp(stats, ('2023-01', '2025-12'))
        self.assertEqual((stats[0]['overlap_start'], stats[0]['overlap_end']),
                         ('2023-01', '2026-07'))
        self.assertEqual((stats[1]['overlap_start'], stats[1]['overlap_end']),
                         ('2023-04', '2026-07'))
        self.assertFalse(stats[2]['has_overlap'])


class TestClampWithContractPipeline(unittest.TestCase):
    """需求1：与合同叠加（v1.1.53）协同——钳制作用于最终重叠层"""

    def test_clamp_after_contract_full_coverage(self):
        """合同完全覆盖参保期：合同叠加不动 → 钳制正常生效"""
        from modules.insurance.core.contract_overlap import apply_contract_to_stats
        roster = [{'name': '张三', 'idcard': '610102199001011234',
                   'contract_status': 'ok', 'contract_error': '',
                   'contract_raw': '2022-01-01~2027-12-31',
                   'contract_periods': [('2022-01', '2027-12')]}]
        stats, _ = calc_all_stats([_make_person()], year_range=('2023-01', '2025-12'))
        apply_contract_to_stats(stats, roster, year_range=('2023-01', '2025-12'))
        apply_stat_range_clamp(stats, ('2023-01', '2025-12'))
        self.assertEqual((stats[0]['overlap_start'], stats[0]['overlap_end']),
                         ('2023-01', '2026-07'))
        self.assertEqual(stats[0]['overlap_months'], 36)

    def test_clamp_after_contract_segments(self):
        """合同分段（间断）：包络起点早于统计开始 → 钳到统计开始，月数按分段实际求和"""
        from modules.insurance.core.contract_overlap import apply_contract_to_stats
        roster = [{'name': '张三', 'idcard': '610102199001011234',
                   'contract_status': 'ok', 'contract_error': '',
                   'contract_raw': '2022-01-01~2023-06-30、2024-01-01~2026-07-31',
                   'contract_periods': [('2022-01', '2023-06'), ('2024-01', '2026-07')]}]
        stats, _ = calc_all_stats([_make_person()], year_range=('2023-01', '2025-12'))
        apply_contract_to_stats(stats, roster, year_range=('2023-01', '2025-12'))
        # 合同包络 = 2022-09~2026-07（首末段包络，首段与参保重叠 2022-09 起相交）
        self.assertEqual(stats[0]['overlap_start'], '2022-09')
        apply_stat_range_clamp(stats, ('2023-01', '2025-12'))
        # 钳制后显示 2023-01~2026-07；月数 = 分段实际（2023:6 + 2024:12 + 2025:12）
        self.assertEqual((stats[0]['overlap_start'], stats[0]['overlap_end']),
                         ('2023-01', '2026-07'))
        self.assertEqual(stats[0]['overlap_months'], 30)
        self.assertEqual(stats[0]['yearly_months'].get(2023), 6)

    def test_year_cols_recompute_after_clamp(self):
        """年度列按钳制后重叠层重算：2022 不出现（钳制剔除）"""
        stats, _ = calc_all_stats([_make_person()], year_range=('2023-01', '2025-12'))
        apply_stat_range_clamp(stats, ('2023-01', '2025-12'))
        from modules.insurance.core.stats_calculator import get_overlap_years
        year_cols = get_overlap_years(
            [ps for ps in stats if ps['has_overlap']], year_range=('2023-01', '2025-12'))
        self.assertEqual(year_cols, [2023, 2024, 2025])


# ==================== 需求5：账号管理删除"用户不存在"修复 ====================
import sqlite3
from unittest import mock

from core import auth as core_auth


def _memory_db(users):
    """构造内存 users 库（模拟本地数据库）"""
    conn = sqlite3.connect(':memory:')
    conn.row_factory = sqlite3.Row
    conn.execute(
        'CREATE TABLE users (id INTEGER PRIMARY KEY, username TEXT, '
        'password_hash TEXT, is_admin INTEGER, is_active INTEGER, created_at TEXT)')
    for u in users:
        conn.execute(
            'INSERT INTO users (id, username, password_hash, is_admin, is_active, created_at) '
            'VALUES (?, ?, ?, ?, ?, ?)',
            (u['id'], u['username'], 'x', u.get('is_admin', 0),
             u.get('is_active', 1), u.get('created_at', '')))
    conn.commit()
    return conn


LOCAL_USERS = [
    {'id': 1, 'username': 'admin', 'is_admin': 1},
    {'id': 2, 'username': 'localuser', 'is_admin': 0},
]
REMOTE_USERS = [
    {'id': 1, 'username': 'admin', 'is_admin': 1, 'is_active': 1,
     'created_at': '2026-07-01 10:00:00'},
    {'id': 5, 'username': '123456', 'is_admin': 0, 'is_active': 1,
     'created_at': '2026-08-03 09:00:00'},
]


class _RequestCtxMixin:
    """提供 Flask 请求上下文（core_auth 的远程分支直接访问 session）"""

    def setUp(self):
        import app as app_module
        app_module.app.config['TESTING'] = True
        self._ctx = app_module.app.test_request_context()
        self._ctx.push()

    def tearDown(self):
        self._ctx.pop()


def _file_db_dir():
    """临时目录（本地模式测试用文件库，连接关闭后状态仍在）"""
    import tempfile
    return tempfile.mkdtemp(prefix='ly_test_users_')


class TestGetUserByIdRemoteAware(_RequestCtxMixin, unittest.TestCase):
    """需求5：get_user_by_id 远程模式走云端列表（修复"用户不存在"误报）"""

    def test_remote_user_found(self):
        """云端存在的账号（如 123456）→ 能查到，不再误报不存在"""
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'users': REMOTE_USERS}, 200)):
            user = core_auth.get_user_by_id(5)
        self.assertIsNotNone(user)
        self.assertEqual(user['username'], '123456')

    def test_remote_user_absent(self):
        """云端列表确认不存在 → 返回 None"""
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'users': REMOTE_USERS}, 200)):
            self.assertIsNone(core_auth.get_user_by_id(99))

    def test_remote_failure_falls_back_local(self):
        """云端不可达（503）→ 回退本地库查询"""
        db = _memory_db(LOCAL_USERS)
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'error': '无法连接认证服务器，请检查网络'}, 503)), \
             mock.patch.object(core_auth, 'get_db', return_value=db):
            user = core_auth.get_user_by_id(2)
        self.assertIsNotNone(user)
        self.assertEqual(user['username'], 'localuser')

    def test_local_mode_unchanged(self):
        """本地模式行为不变：直接查本地库"""
        db = _memory_db(LOCAL_USERS)
        with mock.patch.object(core_auth, '_is_remote', return_value=False), \
             mock.patch.object(core_auth, 'get_db', return_value=db):
            user = core_auth.get_user_by_id(1)
        self.assertEqual(user['username'], 'admin')
        self.assertIsNone(core_auth.get_user_by_id(42))


class TestDeleteUserRemoteSemantics(_RequestCtxMixin, unittest.TestCase):
    """需求5：delete_user 远程模式以云端为准，非200不再假成功回退本地"""

    def test_remote_delete_ok(self):
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'ok': True}, 200)):
            self.assertIsTrue = self.assertTrue
            self.assertIsTrue(core_auth.delete_user(5))

    def test_remote_delete_not_found_returns_error(self):
        """云端返回404 → 返回错误信息字符串，绝不回退本地删除（假成功）"""
        db = _memory_db(LOCAL_USERS)
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'error': '用户不存在'}, 404)), \
             mock.patch.object(core_auth, 'get_db', return_value=db) as gdb:
            result = core_auth.delete_user(5)
        self.assertEqual(result, '用户不存在')
        gdb.assert_not_called()  # 本地库未被触碰

    def test_remote_delete_service_down_no_fake_success(self):
        """云端不可达 → 返回错误信息字符串，不假成功"""
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'error': '无法连接认证服务器，请检查网络'}, 503)):
            result = core_auth.delete_user(5)
        self.assertIsInstance(result, str)

    def test_local_delete_unchanged(self):
        """本地模式删除行为不变（临时文件库，验证真实落盘删除）"""
        import os
        db_dir = _file_db_dir()
        db_path = os.path.join(db_dir, 'users.db')
        seed = _memory_db(LOCAL_USERS)
        # 用文件库重放种子数据（内存库 close 即销毁，无法复用）
        fdb = sqlite3.connect(db_path)
        fdb.execute(
            'CREATE TABLE users (id INTEGER PRIMARY KEY, username TEXT, '
            'password_hash TEXT, is_admin INTEGER, is_active INTEGER, created_at TEXT)')
        for u in LOCAL_USERS:
            fdb.execute(
                'INSERT INTO users (id, username, password_hash, is_admin, is_active, created_at) '
                'VALUES (?, ?, ?, ?, ?, ?)',
                (u['id'], u['username'], 'x', u.get('is_admin', 0),
                 u.get('is_active', 1), ''))
        fdb.commit()
        fdb.close()
        seed.close()
        with mock.patch.object(core_auth, '_is_remote', return_value=False):
            with mock.patch.object(core_auth, 'get_db',
                                   return_value=sqlite3.connect(db_path)):
                self.assertTrue(core_auth.delete_user(2))
        check = sqlite3.connect(db_path)
        remain = [r[0] for r in check.execute(
            'SELECT username FROM users').fetchall()]
        check.close()
        self.assertEqual(remain, ['admin'])


class TestToggleResetRemoteSemantics(_RequestCtxMixin, unittest.TestCase):
    """需求5：toggle/reset_password 远程模式业务错误透传"""

    def test_remote_toggle_ok(self):
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'ok': True, 'is_active': False}, 200)):
            self.assertFalse(core_auth.toggle_user_active(5))

    def test_remote_toggle_error_passthrough(self):
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'error': '用户不存在'}, 404)):
            self.assertEqual(core_auth.toggle_user_active(5), '用户不存在')

    def test_remote_reset_ok(self):
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'ok': True}, 200)):
            self.assertTrue(core_auth.reset_password(5, 'newpass123'))

    def test_remote_reset_error_passthrough(self):
        with mock.patch.object(core_auth, '_is_remote', return_value=True), \
             mock.patch.object(core_auth, '_remote_request',
                               return_value=({'error': '用户不存在'}, 404)):
            self.assertEqual(core_auth.reset_password(5, 'newpass123'), '用户不存在')


class TestUserManagementEndpoints(unittest.TestCase):
    """需求5：app.py 端点级验证——远程账号可删除、错误透传"""

    def _client_with_admin(self):
        import app as app_module
        app_module.app.config['TESTING'] = True
        client = app_module.app.test_client()
        with client.session_transaction() as sess:
            sess['user_id'] = 1
            sess['is_admin'] = True
        return client

    def test_delete_remote_user_success(self):
        """删除云端账号（如 123456, id=5）→ 成功，不再误报不存在"""
        client = self._client_with_admin()
        remote_user = {'id': 5, 'username': '123456', 'is_admin': 0,
                       'is_active': 1, 'created_at': '2026-08-03'}
        with mock.patch.object(core_auth, 'get_user_by_id', return_value=remote_user), \
             mock.patch.object(core_auth, 'delete_user', return_value=True) as m_del:
            resp = client.post('/api/users/5/delete')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['ok'])
        m_del.assert_called_once_with(5)

    def test_delete_business_error_passthrough(self):
        """云端返回业务错误 → 400 + 错误信息透传"""
        client = self._client_with_admin()
        remote_user = {'id': 5, 'username': '123456', 'is_admin': 0, 'is_active': 1}
        with mock.patch.object(core_auth, 'get_user_by_id', return_value=remote_user), \
             mock.patch.object(core_auth, 'delete_user', return_value='用户不存在'):
            resp = client.post('/api/users/5/delete')
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.get_json()['error'], '用户不存在')

    def test_delete_user_truly_absent(self):
        """云端与本地都查不到 → 404 用户不存在"""
        client = self._client_with_admin()
        with mock.patch.object(core_auth, 'get_user_by_id', return_value=None):
            resp = client.post('/api/users/99/delete')
        self.assertEqual(resp.status_code, 404)
        self.assertEqual(resp.get_json()['error'], '用户不存在')

    def test_toggle_remote_user_success(self):
        client = self._client_with_admin()
        remote_user = {'id': 5, 'username': '123456', 'is_admin': 0, 'is_active': 1}
        with mock.patch.object(core_auth, 'get_user_by_id', return_value=remote_user), \
             mock.patch.object(core_auth, 'toggle_user_active', return_value=False):
            resp = client.post('/api/users/5/toggle')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.get_json()['ok'])

    def test_toggle_error_passthrough(self):
        client = self._client_with_admin()
        remote_user = {'id': 5, 'username': '123456', 'is_admin': 0, 'is_active': 1}
        with mock.patch.object(core_auth, 'get_user_by_id', return_value=remote_user), \
             mock.patch.object(core_auth, 'toggle_user_active', return_value='操作失败'):
            resp = client.post('/api/users/5/toggle')
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.get_json()['error'], '操作失败')

    def test_reset_password_error_passthrough(self):
        client = self._client_with_admin()
        remote_user = {'id': 5, 'username': '123456', 'is_admin': 0, 'is_active': 1}
        with mock.patch.object(core_auth, 'get_user_by_id', return_value=remote_user), \
             mock.patch.object(core_auth, 'reset_password', return_value='重置失败'):
            resp = client.post('/api/users/5/reset_password',
                               json={'password': 'newpass123'})
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(resp.get_json()['error'], '重置失败')


# ==================== 需求2/3：前端改动字符串断言 ====================
_APP_JS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       'modules', 'insurance', 'static', 'js', 'app.js')
_STYLE_CSS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                          'modules', 'insurance', 'static', 'css', 'style.css')


class TestFrontendButtonFirstColumn(unittest.TestCase):
    """需求2：修改时间段按钮移至统计表第一列"""

    def setUp(self):
        with open(_APP_JS, encoding='utf-8') as f:
            self.src = f.read()

    def test_header_operation_is_first(self):
        """表头"操作"位于"序号"之前"""
        self.assertIn("'操作',", self.src)
        op_idx = self.src.index("'操作',")
        seq_idx = self.src.index("'序号',")
        self.assertLess(op_idx, seq_idx)
        self.assertNotIn("headers.push('操作');", self.src)

    def test_button_cell_before_seq_cell(self):
        """行内按钮单元格位于序号单元格之前"""
        btn_idx = self.src.index('pe-btn" data-idx=')
        seq_idx = self.src.index("html += '<td>' + (srcIdx + 1) + '</td>';")
        self.assertLess(btn_idx, seq_idx)

    def test_data_idx_binding_kept(self):
        """按钮 data-idx 绑定与重绑定逻辑保留（移列不影响功能）"""
        self.assertIn("document.querySelectorAll('.pe-btn')", self.src)
        self.assertIn("btn.getAttribute('data-idx')", self.src)


class TestFrontendOrganizeSearch(unittest.TestCase):
    """需求3：文件整理结果区实时搜索框"""

    def setUp(self):
        with open(_APP_JS, encoding='utf-8') as f:
            self.src = f.read()

    def test_search_input_present(self):
        """搜索框位于文件整理结果面板顶部"""
        self.assertIn('id="orgSearchInput"', self.src)
        self.assertIn('org-search-bar', self.src)

    def test_filter_function_exists(self):
        """过滤函数存在且已绑定 input 事件"""
        self.assertIn('function filterOrganizeResults()', self.src)
        self.assertIn("addEventListener('input', filterOrganizeResults)", self.src)

    def test_empty_state_hint(self):
        """无匹配项时显示友好提示，不出现空白页"""
        self.assertIn('未找到相关文件', self.src)
        self.assertIn('orgSearchEmptyTip', self.src)

    def test_keyword_preserved_on_rerender(self):
        """重渲染（异常处理后）保留搜索关键字"""
        self.assertIn('organizeSearchKeyword', self.src)

    def test_css_present(self):
        """搜索框样式已定义"""
        with open(_STYLE_CSS, encoding='utf-8') as f:
            css = f.read()
        self.assertIn('.org-search-bar', css)
        self.assertIn('.org-search-input', css)


# ==================== 需求4：导出表格公式保留与打开即重算 ====================
class TestExcelFullCalcOnLoad(unittest.TestCase):
    """需求4：导出文件强制打开即重算（fullCalcOnLoad），公式不丢失不失效"""

    def setUp(self):
        import tempfile
        import shutil
        self.tmpdir = tempfile.mkdtemp(prefix='test_v1155_xl_')

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _assert_full_calc(self, path, expect_formula=True):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        self.assertTrue(wb.calculation.fullCalcOnLoad,
                        f'{path} 缺少 fullCalcOnLoad，打开时公式可能不重算')
        formula_found = False
        for row in wb.active.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_found = True
                    break
        wb.close()
        if expect_formula:
            self.assertTrue(formula_found, f'{path} 中未找到公式单元格')

    def test_main_ledger(self):
        """总台账：fullCalcOnLoad 写入且公式保留"""
        from modules.insurance.core.excel_generator import generate_excel
        from modules.insurance.core.stats_calculator import calc_all_stats
        persons = [{'name': '张三', 'idcard': '610102199001011234',
                    'insurances': {t: ('2022-09', '2026-07') for t in
                                   ['养老保险', '医疗保险', '工伤保险', '失业保险']}}]
        ps_list, year_cols = calc_all_stats(persons, year_range=('2023-01', '2025-12'))
        out = os.path.join(self.tmpdir, 'main.xlsx')
        generate_excel(persons, out, stats=(ps_list, year_cols))
        self._assert_full_calc(out)

    def test_yearly_ledger(self):
        """年度台账：fullCalcOnLoad 写入且金额列公式保留（=月数*标准）"""
        from modules.insurance.core.excel_generator import _generate_yearly_ledger
        from modules.insurance.core.stats_calculator import calc_all_stats
        persons = [{'name': '张三', 'idcard': '610102199001011234',
                    'insurances': {t: ('2022-09', '2026-07') for t in
                                   ['养老保险', '医疗保险', '工伤保险', '失业保险']}}]
        ps_list, _ = calc_all_stats(persons, year_range=('2023-01', '2025-12'))
        classified = [(ps_list[0], '重点群体成员')]
        out = _generate_yearly_ledger(2024, classified, {}, '', self.tmpdir, '0904')
        self.assertIsNotNone(out)
        path = out['filepath']
        self.assertTrue(os.path.exists(path))
        self._assert_full_calc(path)


if __name__ == '__main__':
    unittest.main()
