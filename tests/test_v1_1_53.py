# -*- coding: utf-8 -*-
"""v1.1.53 劳动合同起止时间 × 四险参保时间段叠加比对测试

需求（用户确认版）：
1. 花名册新增"劳动合同起止时间"列，上传解析时做日期格式校验并妥善处理缺失
2. 四险参保时间段统计完成后，与合同起止时间叠加比对，重叠部分作为有效参保期
3. 同一员工多段参保区间/多份合同期限，分别计算并合并重叠结果
4. 缺失/异常策略：不裁剪 + 操作记录标注待补（缺失=信息不全，不静默清零）
5. 其余统计规则、计算流程、输出格式均保持不变（仅后处理 calc_all_stats 的重叠层）
"""
import os
import shutil
import sys
import tempfile
import unittest
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.insurance.core.contract_overlap import (
    OPEN_END, parse_contract_cell, intersect_periods, apply_contract_to_stats)
from modules.insurance.core.roster_parser import (
    _extract_roster_from_rows, _stringify_contract_cell, _parse_excel)
from modules.insurance.core.stats_calculator import calc_all_stats
from modules.insurance.core.excel_generator import generate_excel
from modules.insurance import blueprint as bp

ID_ZHANG = '11010519491231002X'  # 校验码合法
ID_LI = '110101199003077758'


def _mk_person(name, idcard, start, end, ins_types=None):
    """构造 group_by_person 产出格式的人员 dict（四险同段）"""
    ins_types = ins_types or ['养老保险', '医疗保险', '工伤保险', '失业保险']
    return {'name': name, 'idcard': idcard,
            'insurances': {t: (start, end) for t in ins_types}}


def _mk_ps(start, end, name='张三', idcard=ID_ZHANG, year_range=None, ins_types=None):
    """用真实 calc_all_stats 造 person_stats（保证字段结构与线上一致）"""
    persons = [_mk_person(name, idcard, start, end, ins_types)]
    ps_list, _ = calc_all_stats(persons, year_range=year_range)
    return ps_list


def _roster_entry(name, idcard=ID_ZHANG, periods=None, status='ok', error=''):
    return {'seq': 1, 'name': name, 'idcard': idcard, 'identity_type': '脱贫人口',
            'contract_periods': periods or [], 'contract_status': status,
            'contract_raw': '', 'contract_error': error}


# ==================== 1. 合同单元格解析 ====================
class TestParseContractCell(unittest.TestCase):

    def test_day_precision_formats(self):
        for text in ['2023年1月5日至2025年12月31日',
                     '2023-01-05~2025-12-31',
                     '2023/1/5-2025/12/31',
                     '2023.1.5至2025.12.31',
                     '20230105-20251231']:
            r = parse_contract_cell(text)
            self.assertEqual(r['status'], 'ok', text)
            self.assertEqual(r['periods'], [('2023-01', '2025-12')], text)

    def test_month_precision_formats(self):
        for text in ['2023年1月-2025年12月', '2023-01~2025-12', '202301-202512']:
            r = parse_contract_cell(text)
            self.assertEqual(r['status'], 'ok', text)
            self.assertEqual(r['periods'], [('2023-01', '2025-12')], text)

    def test_day_counts_whole_month(self):
        """月粒度：起止日所在月份即视为覆盖月（与参保月粒度对齐）"""
        r = parse_contract_cell('2023-01-31~2025-12-01')
        self.assertEqual(r['periods'], [('2023-01', '2025-12')])

    def test_multi_segments_kept(self):
        r = parse_contract_cell('2020-01-01~2021-12-31；2023-03-01~2025-06-30')
        self.assertEqual(r['status'], 'ok')
        self.assertEqual(r['periods'], [('2020-01', '2021-12'), ('2023-03', '2025-06')])

    def test_adjacent_segments_merged(self):
        r = parse_contract_cell('2020-01~2021-12；2022-01~2023-06')
        self.assertEqual(r['periods'], [('2020-01', '2023-06')])

    def test_multi_segments_separator_variants(self):
        for text in ['2020-01~2021-12、2023-01~2025-06',
                     '2020-01~2021-12，2023-01~2025-06',
                     '2020-01~2021-12\n2023-01~2025-06']:
            r = parse_contract_cell(text)
            self.assertEqual(r['periods'],
                             [('2020-01', '2021-12'), ('2023-01', '2025-06')], text)

    def test_open_end_keywords(self):
        for text in ['2023-01-01 至今', '2023年1月 无固定期限', '2023-01 长期']:
            r = parse_contract_cell(text)
            self.assertEqual(r['status'], 'ok', text)
            self.assertEqual(r['periods'], [('2023-01', OPEN_END)], text)

    def test_invalid_day(self):
        r = parse_contract_cell('2023-02-30~2025-12-31')
        self.assertEqual(r['status'], 'invalid')
        self.assertIn('非法日期', r['error'])

    def test_start_after_end(self):
        r = parse_contract_cell('2025-01~2023-12')
        self.assertEqual(r['status'], 'invalid')
        self.assertIn('开始晚于结束', r['error'])

    def test_single_date_without_keyword(self):
        r = parse_contract_cell('2023-01')
        self.assertEqual(r['status'], 'invalid')
        self.assertIn('成对', r['error'])

    def test_garbage_text(self):
        r = parse_contract_cell('暂无合同信息')
        self.assertEqual(r['status'], 'invalid')
        self.assertIn('未识别到日期', r['error'])

    def test_three_dates_in_one_segment(self):
        r = parse_contract_cell('2023-01-01~2024-01-01~2025-01-01')
        self.assertEqual(r['status'], 'invalid')
        self.assertIn('3 个以上', r['error'])

    def test_missing_variants(self):
        for text in [None, '', '   ', '-', '—', '/', '无', 'nan', 'None']:
            r = parse_contract_cell(text)
            self.assertEqual(r['status'], 'missing', repr(text))
            self.assertEqual(r['periods'], [])


# ==================== 2. 花名册合同列解析 ====================
class TestRosterContractColumn(unittest.TestCase):

    def _rows(self, header, cells):
        return [['序号', '姓名', '身份证号', header],
                [1, '张三', ID_ZHANG, cells]]

    def test_header_variants(self):
        for header in ['劳动合同起止时间', '合同期限', '合同起止', '劳动合同', '合同时间']:
            roster = _extract_roster_from_rows(self._rows(header, '2023-01~2025-12'))
            self.assertEqual(len(roster), 1, header)
            self.assertEqual(roster[0]['contract_status'], 'ok', header)
            self.assertEqual(roster[0]['contract_periods'], [('2023-01', '2025-12')], header)

    def test_no_contract_column(self):
        rows = [['序号', '姓名', '身份证号'], [1, '张三', ID_ZHANG]]
        roster = _extract_roster_from_rows(rows)
        self.assertEqual(roster[0]['contract_status'], 'missing')
        self.assertEqual(roster[0]['contract_periods'], [])

    def test_empty_cell(self):
        roster = _extract_roster_from_rows(self._rows('劳动合同起止时间', None))
        self.assertEqual(roster[0]['contract_status'], 'missing')

    def test_invalid_cell_keeps_error(self):
        roster = _extract_roster_from_rows(self._rows('劳动合同起止时间', '随便写'))
        self.assertEqual(roster[0]['contract_status'], 'invalid')
        self.assertTrue(roster[0]['contract_error'])
        self.assertEqual(roster[0]['contract_raw'], '随便写')

    def test_stringify_datetime_and_float_artifact(self):
        self.assertEqual(_stringify_contract_cell(datetime(2023, 1, 5)), '2023-01-05')
        self.assertEqual(_stringify_contract_cell('20230105.0'), '20230105')
        self.assertEqual(_stringify_contract_cell(None), '')

    def test_xlsx_roundtrip(self):
        """真实 xlsx 文件端到端：合同列文本 → 解析为月粒度段"""
        from openpyxl import Workbook
        tmpdir = tempfile.mkdtemp(prefix='test_v1153_xlsx_')
        try:
            path = os.path.join(tmpdir, 'roster.xlsx')
            wb = Workbook()
            ws = wb.active
            ws.append(['序号', '姓名', '身份证号', '劳动合同起止时间'])
            ws.append([1, '张三', ID_ZHANG, '2023-01-01~2025-12-31'])
            ws.append([2, '李四', ID_LI, '2020年3月至今'])
            wb.save(path)
            roster = _parse_excel(path)
            self.assertEqual(len(roster), 2)
            self.assertEqual(roster[0]['contract_periods'], [('2023-01', '2025-12')])
            self.assertEqual(roster[1]['contract_periods'], [('2020-03', OPEN_END)])
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)


# ==================== 3. 区间求交 ====================
class TestIntersectPeriods(unittest.TestCase):

    def test_clip_basic(self):
        self.assertEqual(
            intersect_periods('2023-01', '2025-12', [('2024-01', '2024-12')]),
            [('2024-01', '2024-12')])

    def test_multi_segment(self):
        self.assertEqual(
            intersect_periods('2020-01', '2025-12',
                              [('2020-06', '2021-06'), ('2023-01', '2024-12')]),
            [('2020-06', '2021-06'), ('2023-01', '2024-12')])

    def test_adjacent_segments_merged(self):
        self.assertEqual(
            intersect_periods('2020-01', '2026-12',
                              [('2023-01', '2023-12'), ('2024-01', '2025-12')]),
            [('2023-01', '2025-12')])

    def test_no_overlap(self):
        self.assertEqual(
            intersect_periods('2023-01', '2025-12', [('2026-01', '2026-12')]), [])

    def test_contract_covers_all(self):
        self.assertEqual(
            intersect_periods('2023-01', '2025-12', [('2020-01', '2030-12')]),
            [('2023-01', '2025-12')])

    def test_open_end(self):
        self.assertEqual(
            intersect_periods('2023-01', '2025-12', [('2024-06', OPEN_END)]),
            [('2024-06', '2025-12')])

    def test_base_inside_contract(self):
        self.assertEqual(
            intersect_periods('2020-01', '2030-12', [('2023-01', '2025-12')]),
            [('2023-01', '2025-12')])


# ==================== 4. 叠加比对（统计层裁剪） ====================
class TestApplyContractToStats(unittest.TestCase):

    def test_clip_basic(self):
        ps_list = _mk_ps('2023-01', '2025-12')
        roster = [_roster_entry('张三', periods=[('2024-01', '2024-12')])]
        notes = apply_contract_to_stats(ps_list, roster)
        ps = ps_list[0]
        self.assertEqual(ps['overlap_start'], '2024-01')
        self.assertEqual(ps['overlap_end'], '2024-12')
        self.assertEqual(ps['overlap_months'], 12)
        self.assertEqual(ps['yearly_months'], {2024: 12})
        self.assertEqual(ps['years'], [2024])
        self.assertTrue(ps['has_overlap'])
        self.assertEqual(len(notes), 1)
        self.assertEqual(notes[0]['action'], '合同比对')
        self.assertEqual(notes[0]['old'], '2023-01~2025-12')
        # 各险种原始时间段不动
        self.assertEqual(ps['insurances']['养老保险'], ('2023-01', '2025-12'))

    def test_full_cover_untouched_no_note(self):
        ps_list = _mk_ps('2023-01', '2025-12')
        roster = [_roster_entry('张三', periods=[('2020-01', '2030-12')])]
        notes = apply_contract_to_stats(ps_list, roster)
        self.assertEqual(ps_list[0]['overlap_start'], '2023-01')
        self.assertEqual(ps_list[0]['overlap_months'], 36)
        self.assertEqual(notes, [])

    def test_missing_not_clipped_with_note(self):
        ps_list = _mk_ps('2023-01', '2025-12')
        roster = [_roster_entry('张三', status='missing')]
        notes = apply_contract_to_stats(ps_list, roster)
        self.assertEqual(ps_list[0]['overlap_start'], '2023-01')  # 不裁剪
        self.assertEqual(ps_list[0]['overlap_months'], 36)
        self.assertEqual(len(notes), 1)
        self.assertIn('未登记', notes[0]['new'])

    def test_invalid_not_clipped_with_note(self):
        ps_list = _mk_ps('2023-01', '2025-12')
        roster = [_roster_entry('张三', status='invalid', error='非法日期: 2023-02-30')]
        notes = apply_contract_to_stats(ps_list, roster)
        self.assertEqual(ps_list[0]['overlap_months'], 36)  # 不裁剪
        self.assertEqual(len(notes), 1)
        self.assertIn('格式异常', notes[0]['new'])

    def test_no_overlap_empties_result(self):
        ps_list = _mk_ps('2023-01', '2025-12')
        roster = [_roster_entry('张三', periods=[('2026-01', '2026-12')])]
        notes = apply_contract_to_stats(ps_list, roster)
        ps = ps_list[0]
        self.assertFalse(ps['has_overlap'])
        self.assertIsNone(ps['overlap_start'])
        self.assertIsNone(ps['overlap_end'])
        self.assertEqual(ps['overlap_months'], 0)
        self.assertEqual(ps['yearly_months'], {})
        self.assertEqual(ps['years'], [])
        self.assertEqual(len(notes), 1)
        self.assertIn('无重叠', notes[0]['new'])

    def test_multi_segment_envelope_with_accurate_months(self):
        """多段合同：包络为首末段，月数按分段实际求和（间断月不计入）"""
        ps_list = _mk_ps('2020-01', '2025-12')
        roster = [_roster_entry('张三', periods=[('2020-06', '2021-06'),
                                                ('2023-01', '2024-06')])]
        notes = apply_contract_to_stats(ps_list, roster)
        ps = ps_list[0]
        self.assertEqual(ps['overlap_start'], '2020-06')  # 首段开始
        self.assertEqual(ps['overlap_end'], '2024-06')    # 末段结束
        self.assertEqual(ps['overlap_months'], 7 + 6 + 12 + 6)  # 31，间断月不计
        self.assertEqual(ps['yearly_months'],
                         {2020: 7, 2021: 6, 2022: 0, 2023: 12, 2024: 6})
        self.assertEqual(len(notes), 1)
        self.assertIn('分段', notes[0]['new'])

    def test_year_range_filter(self):
        """year_range 筛选：年度月数与总月数跟筛选走（v1.1.38 一致性）"""
        ps_list = _mk_ps('2023-01', '2025-12', year_range=('2024-01', '2024-12'))
        roster = [_roster_entry('张三', periods=[('2023-06', '2025-06')])]
        notes = apply_contract_to_stats(ps_list, roster, year_range=('2024-01', '2024-12'))
        ps = ps_list[0]
        self.assertEqual(ps['overlap_start'], '2023-06')
        self.assertEqual(ps['overlap_end'], '2025-06')
        self.assertEqual(ps['yearly_months'], {2023: 0, 2024: 12, 2025: 0})
        self.assertEqual(ps['overlap_months'], 12)
        self.assertEqual(len(notes), 1)

    def test_not_in_roster_untouched(self):
        ps_list = _mk_ps('2023-01', '2025-12')
        roster = [_roster_entry('李四', idcard=ID_LI, periods=[('2024-01', '2024-12')])]
        notes = apply_contract_to_stats(ps_list, roster)
        self.assertEqual(ps_list[0]['overlap_months'], 36)
        self.assertEqual(notes, [])

    def test_no_insurance_overlap_skipped(self):
        """四险本身无重叠（仅 1 个险种有数据）→ 无需比对，不标注"""
        ps_list = _mk_ps('2023-01', '2025-12', ins_types=['养老保险'])
        self.assertFalse(ps_list[0]['has_overlap'])
        roster = [_roster_entry('张三', periods=[('2024-01', '2024-12')])]
        notes = apply_contract_to_stats(ps_list, roster)
        self.assertFalse(ps_list[0]['has_overlap'])
        self.assertEqual(notes, [])

    def test_empty_roster_noop(self):
        ps_list = _mk_ps('2023-01', '2025-12')
        self.assertEqual(apply_contract_to_stats(ps_list, []), [])
        self.assertEqual(ps_list[0]['overlap_months'], 36)


# ==================== 5. Excel 生成器 stats 参数 ====================
class TestGenerateExcelStatsParam(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp(prefix='test_v1153_xl_')

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _read_row3(self, path):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        row = [ws.cell(row=3, column=c).value for c in range(1, 13)]
        wb.close()
        return row

    def test_stats_param_used_when_given(self):
        """传入已裁剪 stats → Excel 用传入值（与页面 JSON 一致），不再内部重算"""
        persons = [_mk_person('张三', ID_ZHANG, '2023-01', '2025-12')]
        ps_list, _ = calc_all_stats(persons)
        # 模拟合同裁剪后的统计（篡改重叠层）
        ps_list[0].update({'overlap_start': '2024-01', 'overlap_end': '2024-12',
                           'overlap_months': 12, 'years': [2024],
                           'yearly_months': {2024: 12}})
        out = os.path.join(self.tmpdir, 'with_stats.xlsx')
        generate_excel(persons, out, stats=(ps_list, [2024]))
        row = self._read_row3(out)
        # v1.1.54：第5列新增"劳动合同起止时间"，后续列 +1
        self.assertEqual(row[4], '-')                 # 第5列：合同列（无花名册→'-'）
        self.assertEqual(row[9], '2024-01~2024-12')   # 第10列：重叠时间段
        self.assertEqual(row[10], 12)                 # 第11列：重叠月数
        self.assertEqual(row[11], 12)                 # 第12列：2024年度月数

    def test_default_behavior_without_stats(self):
        """不传 stats → 保持原行为（内部自行统计）"""
        persons = [_mk_person('张三', ID_ZHANG, '2023-01', '2025-12')]
        out = os.path.join(self.tmpdir, 'no_stats.xlsx')
        generate_excel(persons, out)
        row = self._read_row3(out)
        self.assertEqual(row[9], '2023-01~2025-12')   # 第10列：重叠时间段（v1.1.54 +1）
        self.assertEqual(row[10], 36)                 # 第11列：重叠月数


# ==================== 6. _rebuild_result 全链路集成 ====================
class TestRebuildResultContractIntegration(unittest.TestCase):
    """_rebuild_result 漏斗：合同裁剪 → 公开 result / operation_log / Excel 三处一致"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp(prefix='test_v1153_bp_')
        self.task_id = 'testv1153'
        self._old_out = bp.OUTPUT_DIR
        bp.OUTPUT_DIR = os.path.join(self.tmpdir, 'outputs')
        os.makedirs(bp.OUTPUT_DIR, exist_ok=True)

    def tearDown(self):
        with bp.tasks_lock:
            bp.tasks.pop(self.task_id, None)
        bp.OUTPUT_DIR = self._old_out
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _mkimg(self, name):
        path = os.path.join(self.tmpdir, name)
        with open(path, 'wb') as f:
            f.write(b'\xff\xd8fakeimg')
        return path

    def _seed_task(self, roster):
        success = []
        for i, t in enumerate(['养老保险', '医疗保险', '工伤保险', '失业保险']):
            success.append({
                'filename': f'img{i}.jpg', 'name': '张三', 'idcard': ID_ZHANG,
                'insurance_type': t, 'period': ('2023-01', '2025-12'),
                'company_name': '鲁岳测试公司', 'raw_text': '', 'error': None,
                '_source_path': self._mkimg(f'img{i}.jpg'), '_source_origin': f'img{i}.jpg'})
        with bp.tasks_lock:
            bp.tasks[self.task_id] = {
                'status': 'done', 'current': 4, 'total': 4,
                'message': '处理完成', 'files': [], 'created_at': '',
                'paused': False, 'cancelled': False,
                'result': {
                    '_success_results': success,
                    '_excluded_results': [],
                    '_failed_results': [],
                    '_all_files': [f'img{i}.jpg' for i in range(4)],
                    '_task_dir': self.tmpdir,
                    '_year_range': None,
                    '_roster': roster,
                    '_roster_company': '鲁岳测试公司',
                    '_roster_source_path': '',
                    '_company_name': '鲁岳测试公司',
                    '_ocr_companies': {'鲁岳测试公司': 4},
                    '_company_mismatch_files': [],
                    '_period_overrides': {},
                    '_manual_log': [],
                },
            }

    def test_clipped_end_to_end(self):
        """合同 2024 全年：重叠层被裁剪为 2024-01~2024-12，三处输出一致"""
        self._seed_task([_roster_entry('张三', periods=[('2024-01', '2024-12')])])
        result = bp._rebuild_result(self.task_id)
        ps = result['person_stats'][0]
        self.assertEqual(ps['overlap_start'], '2024-01')
        self.assertEqual(ps['overlap_end'], '2024-12')
        self.assertEqual(ps['overlap_months'], 12)
        self.assertEqual(result['year_cols'], [2024])
        # 操作记录含合同比对提示（不落盘、每次重建重新生成）
        logs = [l for l in result['operation_log'] if l['action'] == '合同比对']
        self.assertEqual(len(logs), 1)
        self.assertEqual(logs[0]['old'], '2023-01~2025-12')
        # Excel 与 JSON 一致（stats 参数传入）
        from openpyxl import load_workbook
        wb = load_workbook(result['excel_path'])
        ws = wb.active
        # v1.1.54：合同列插入后，重叠时间段/月数列顺移至 10/11
        self.assertEqual(ws.cell(row=3, column=10).value, '2024-01~2024-12')
        self.assertEqual(ws.cell(row=3, column=11).value, 12)
        wb.close()
        # 内部状态里 person_stats 不带合同字段泄漏（公开字段裁剪正确即可）
        self.assertNotIn('_manual_log', result)

    def test_missing_contract_keeps_stats_with_note(self):
        """合同列缺失：不裁剪 + 标注待补"""
        self._seed_task([_roster_entry('张三', status='missing')])
        result = bp._rebuild_result(self.task_id)
        ps = result['person_stats'][0]
        self.assertEqual(ps['overlap_start'], '2023-01')  # 未裁剪
        self.assertEqual(ps['overlap_months'], 36)
        self.assertEqual(result['year_cols'], [2023, 2024, 2025])
        logs = [l for l in result['operation_log'] if l['action'] == '合同比对']
        self.assertEqual(len(logs), 1)
        self.assertIn('未登记', logs[0]['new'])

    def test_no_roster_no_notes(self):
        """无花名册：无任何合同提示，统计原样"""
        self._seed_task([])
        result = bp._rebuild_result(self.task_id)
        self.assertEqual(result['person_stats'][0]['overlap_months'], 36)
        self.assertEqual(result['operation_log'], [])


if __name__ == '__main__':
    unittest.main(verbosity=2)
