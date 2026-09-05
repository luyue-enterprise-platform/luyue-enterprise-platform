# -*- coding: utf-8 -*-
"""v1.1.54 测试套件

需求1（中断信息明细统计规则，此前仅分析未落码）：
    有中断信息明细时，统计参保开始时间 = 最后一段中断的结束月 +1 月；
    多段中断取结束时间最晚的一段；无中断明细/区块为空时维持现有规则不变。

需求2（劳动合同起止时间列展示）：预览列表与统计表新增"劳动合同起止时间"列，
    位于养老保险参保证明时间段之前（相关用例在本文件 TestContractColumn* 中，
    与需求2实现一并补充）。
"""
import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules.insurance.core.data_parser import (
    _extract_interruption_periods,
    _ym_plus_one,
    apply_interruption_start_rule,
    get_full_period,
    get_full_period_from_items,
)


def _mk_item(text, x, y):
    return {'text': text, 'x': x, 'y': y, 'score': 0.99}


# 城镇职工医保年度表格基座（2021-05~2026-07，同 v1.1.51 基线）
MEDICAL_HEADER = '''城镇职工基本医疗保险参保缴费证明
姓名：张明涛 身份证号：610502199001011234
缴费年度 缴费月份 缴费年度 缴费月份 缴费年度 缴费月份
2016 2（月） 2021 8（月） 2024 12（月）
2017 12（月） 2022 12（月） 2025 12（月）
2018 4（月） 2023 12（月） 2026 7（月）
'''


class TestExtractInterruptionPeriods(unittest.TestCase):
    """中断时间段提取单元测试"""

    def test_no_interruption_returns_empty(self):
        self.assertEqual(_extract_interruption_periods(MEDICAL_HEADER), [])

    def test_same_line_six_digit(self):
        text = MEDICAL_HEADER + '中断起止时间 201805-202104\n'
        self.assertEqual(_extract_interruption_periods(text), [('2018-05', '2021-04')])

    def test_chinese_format(self):
        text = MEDICAL_HEADER + '中断起止时间：2018年05月至2021年04月\n'
        self.assertEqual(_extract_interruption_periods(text), [('2018-05', '2021-04')])

    def test_dot_format(self):
        text = MEDICAL_HEADER + '中断时间 2018.05-2021.04\n'
        self.assertEqual(_extract_interruption_periods(text), [('2018-05', '2021-04')])

    def test_data_on_next_line(self):
        text = MEDICAL_HEADER + '中断信息明细\n中断起止时间\n201805-202104\n'
        self.assertEqual(_extract_interruption_periods(text), [('2018-05', '2021-04')])

    def test_multiple_segments(self):
        text = MEDICAL_HEADER + '中断信息明细\n201805-202104\n202301-202409\n'
        self.assertEqual(_extract_interruption_periods(text),
                         [('2018-05', '2021-04'), ('2023-01', '2024-09')])

    def test_header_only_no_dates(self):
        """只有"中断信息明细"表头、无日期数据 → 空（维持现有规则）"""
        text = MEDICAL_HEADER + '中断信息明细\n本证明仅作参考\n'
        self.assertEqual(_extract_interruption_periods(text), [])

    def test_periods_outside_zone_not_captured(self):
        """中断区块之外的日期不得被采集中断段"""
        text = MEDICAL_HEADER + '中断信息明细\n201805-202104\n打印时间 2026年08月\n'
        self.assertEqual(_extract_interruption_periods(text), [('2018-05', '2021-04')])


class TestYmPlusOne(unittest.TestCase):

    def test_normal_month(self):
        self.assertEqual(_ym_plus_one('2024-09'), '2024-10')

    def test_december_rollover(self):
        self.assertEqual(_ym_plus_one('2023-12'), '2024-01')


class TestApplyInterruptionStartRule(unittest.TestCase):
    """中断统计开始时间规则单元测试"""

    def test_no_interruption_period_unchanged(self):
        self.assertEqual(apply_interruption_start_rule(MEDICAL_HEADER, ('2021-05', '2026-07')),
                         ('2021-05', '2026-07'))

    def test_single_segment_start_overridden(self):
        """张明涛案例：中断 202305-202409 → 统计开始 2024-10"""
        text = MEDICAL_HEADER + '中断起止时间 202305-202409\n'
        self.assertEqual(apply_interruption_start_rule(text, ('2021-05', '2026-07')),
                         ('2024-10', '2026-07'))

    def test_multiple_segments_uses_latest_end(self):
        text = MEDICAL_HEADER + '中断信息明细\n201805-202104\n202301-202409\n'
        self.assertEqual(apply_interruption_start_rule(text, ('2021-05', '2026-07')),
                         ('2024-10', '2026-07'))

    def test_december_end_rollover(self):
        text = MEDICAL_HEADER + '中断起止时间 202301-202312\n'
        self.assertEqual(apply_interruption_start_rule(text, ('2021-05', '2026-07')),
                         ('2024-01', '2026-07'))

    def test_none_period_stays_none(self):
        text = MEDICAL_HEADER + '中断起止时间 201805-202104\n'
        self.assertIsNone(apply_interruption_start_rule(text, None))

    def test_start_beyond_end_returns_none(self):
        """中断结束月+1 超过参保结束月 → 无有效参保段"""
        text = MEDICAL_HEADER + '中断起止时间 202305-202612\n'
        self.assertIsNone(apply_interruption_start_rule(text, ('2021-05', '2026-07')))

    def test_start_equal_end_kept(self):
        text = MEDICAL_HEADER + '中断起止时间 202305-202606\n'
        self.assertEqual(apply_interruption_start_rule(text, ('2021-05', '2026-07')),
                         ('2026-07', '2026-07'))


class TestInterruptionRuleEndToEnd(unittest.TestCase):
    """经由 get_full_period / get_full_period_from_items 的端到端验证"""

    def test_text_path_single_segment(self):
        """张明涛案例（文本路径）：统计开始时间被改写为中断结束月+1"""
        text = MEDICAL_HEADER + '中断信息明细\n中断起止时间 202305-202409\n'
        self.assertEqual(get_full_period(text), ('2024-10', '2026-07'))

    def test_text_path_multiple_segments(self):
        text = MEDICAL_HEADER + '中断信息明细\n201805-202104\n202301-202409\n'
        self.assertEqual(get_full_period(text), ('2024-10', '2026-07'))

    def test_text_path_chinese_format(self):
        text = MEDICAL_HEADER + '中断起止时间：2023年05月至2024年09月\n'
        self.assertEqual(get_full_period(text), ('2024-10', '2026-07'))

    def test_text_path_no_interruption_baseline(self):
        """无中断明细 → 现有规则不变（回归）"""
        self.assertEqual(get_full_period(MEDICAL_HEADER), ('2021-05', '2026-07'))

    def test_items_path_single_segment(self):
        """生产路径 parse_ocr_result_from_image → get_full_period_from_items"""
        items = []
        xs = [100, 200, 400, 500, 700, 800]
        for i, x in enumerate(xs):
            items.append(_mk_item('缴费年度' if i % 2 == 0 else '缴费月份', x, 100))
        rows = [
            [('2016', 100), ('2（月）', 200), ('2021', 400), ('8（月）', 500), ('2024', 700), ('12（月）', 800)],
            [('2017', 100), ('12（月）', 200), ('2022', 400), ('12（月）', 500), ('2025', 700), ('12（月）', 800)],
            [('2018', 100), ('4（月）', 200), ('2023', 400), ('12（月）', 500), ('2026', 700), ('7（月）', 800)],
        ]
        for ri, row in enumerate(rows):
            for text, x in row:
                items.append(_mk_item(text, x, 150 + ri * 30))
        items.append(_mk_item('中断信息明细', 100, 300))
        items.append(_mk_item('中断起止时间', 100, 330))
        items.append(_mk_item('202305-202409', 300, 330))
        self.assertEqual(get_full_period_from_items(items), ('2024-10', '2026-07'))

    def test_v1151_baseline_still_passes(self):
        """v1.1.51 基线：中断 201805-202104，结束月+1=2021-05 与解析起点一致"""
        text = MEDICAL_HEADER + '中断起止时间 201805-202104\n'
        self.assertEqual(get_full_period(text), ('2021-05', '2026-07'))


# ==================== 需求2：劳动合同起止时间列展示 ====================
import shutil
import tempfile

from modules.insurance.core.contract_overlap import contract_display_text
from modules.insurance.core.excel_generator import generate_excel
from modules.insurance.core.stats_calculator import calc_all_stats
from modules.insurance import blueprint as bp

ID_ZHANG = '11010519491231002X'


def _mk_person(name, idcard, start, end):
    return {'name': name, 'idcard': idcard,
            'insurances': {t: (start, end) for t in
                           ['养老保险', '医疗保险', '工伤保险', '失业保险']}}


def _roster_entry(name, idcard=ID_ZHANG, raw='', identity_type='脱贫人口'):
    return {'seq': 1, 'name': name, 'idcard': idcard, 'identity_type': identity_type,
            'contract_periods': [], 'contract_status': 'ok' if raw else 'missing',
            'contract_raw': raw, 'contract_error': ''}


class TestContractDisplayText(unittest.TestCase):
    """花名册条目 → 合同起止时间展示文本"""

    def test_none_entry(self):
        self.assertEqual(contract_display_text(None), '-')

    def test_empty_raw(self):
        self.assertEqual(contract_display_text(_roster_entry('张三', raw='')), '-')

    def test_raw_returned_verbatim(self):
        for raw in ['2023-01-05~2025-12-31', '2023-01 至今', '2023年1月 无固定期限']:
            self.assertEqual(contract_display_text(_roster_entry('张三', raw=raw)), raw)

    def test_raw_whitespace_stripped(self):
        self.assertEqual(contract_display_text(_roster_entry('张三', raw='  2023-01~2025-12  ')),
                         '2023-01~2025-12')


class TestBlueprintContractField(unittest.TestCase):
    """blueprint 公开 person_stats 的 contract 字段"""

    def _index(self, roster):
        return bp._build_roster_index(roster)

    def test_match_by_idcard(self):
        roster = [_roster_entry('张三', raw='2023-01-05~2025-12-31')]
        person = {'name': '张三', 'idcard': ID_ZHANG}
        self.assertEqual(bp._get_contract_display(person, self._index(roster)),
                         '2023-01-05~2025-12-31')

    def test_match_by_name_fallback(self):
        roster = [_roster_entry('张三', idcard='', raw='2023-01 至今')]
        person = {'name': '张三', 'idcard': '999'}
        self.assertEqual(bp._get_contract_display(person, self._index(roster)), '2023-01 至今')

    def test_no_match_returns_dash(self):
        person = {'name': '李四', 'idcard': '888'}
        self.assertEqual(bp._get_contract_display(person, self._index([])), '-')


class TestExcelContractColumn(unittest.TestCase):
    """总台账/年度台账 Excel 合同列（位于参保证明时间段列之前）"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp(prefix='test_v1154_xl_')

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _gen(self, roster):
        persons = [_mk_person('张三', ID_ZHANG, '2023-01', '2025-12')]
        out = os.path.join(self.tmpdir, 'ledger.xlsx')
        result = generate_excel(persons, out, roster=roster)
        return out, result

    def test_main_ledger_header_and_value(self):
        """非退役版：第5列=劳动合同起止时间，第6列=养老保险参保证明时间段"""
        from openpyxl import load_workbook
        out, _ = self._gen([_roster_entry('张三', raw='2023-01-05~2025-12-31')])
        wb = load_workbook(out)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=5).value, '劳动合同起止时间')
        self.assertEqual(ws.cell(row=2, column=6).value, '养老保险参保证明时间段')
        self.assertEqual(ws.cell(row=3, column=5).value, '2023-01-05~2025-12-31')
        self.assertEqual(ws.cell(row=3, column=6).value, '2023-01~2025-12')  # 养老列内容未串位
        wb.close()

    def test_main_ledger_no_contract_shows_dash(self):
        from openpyxl import load_workbook
        out, _ = self._gen([_roster_entry('张三', raw='')])
        wb = load_workbook(out)
        self.assertEqual(wb.active.cell(row=3, column=5).value, '-')
        wb.close()

    def test_main_ledger_tuiwu_variant(self):
        """退役版：第7列=劳动合同起止时间，第8列=养老保险参保证明时间段"""
        from openpyxl import load_workbook
        out, _ = self._gen([_roster_entry('张三', raw='2023-01 至今',
                                          identity_type='自主就业退役士兵')])
        wb = load_workbook(out)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=7).value, '劳动合同起止时间')
        self.assertEqual(ws.cell(row=2, column=8).value, '养老保险参保证明时间段')
        self.assertEqual(ws.cell(row=3, column=7).value, '2023-01 至今')
        wb.close()

    def test_yearly_ledger_contract_column(self):
        """年度台账：合同列位于"本年度参保证明时间段"之前"""
        from openpyxl import load_workbook
        _, result = self._gen([_roster_entry('张三', raw='2023-01-05~2025-12-31')])
        yearly = result.get('yearly_ledger_files', [])
        self.assertTrue(yearly, '应生成年度台账')
        wb = load_workbook(yearly[0]['filepath'])
        ws = wb.active
        self.assertEqual(ws.cell(row=3, column=5).value, '劳动合同起止时间')
        self.assertIn('本年度参保证明时间段', ws.cell(row=3, column=6).value)
        self.assertEqual(ws.cell(row=4, column=5).value, '2023-01-05~2025-12-31')
        wb.close()


class TestFrontendContractColumn(unittest.TestCase):
    """前端预览表：表头含劳动合同起止时间且在养老保险列之前 + 渲染 ps.contract"""

    def test_app_js_contract_column(self):
        app_js = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              'modules', 'insurance', 'static', 'js', 'app.js')
        with open(app_js, 'r', encoding='utf-8') as f:
            src = f.read()
        self.assertIn("'劳动合同起止时间',", src)
        self.assertIn('ps.contract', src)
        # 表头顺序：劳动合同起止时间在养老保险参保证明时间段之前
        h_contract = src.index("'劳动合同起止时间',")
        h_pension = src.index("'养老保险参保证明时间段'")
        self.assertLess(h_contract, h_pension)


if __name__ == '__main__':
    unittest.main()
